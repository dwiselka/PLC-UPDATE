import paramiko
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import threading
import os
import time
import socket
import subprocess
from datetime import datetime
import pytz
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font
import queue
from contextlib import contextmanager
from concurrent.futures import ThreadPoolExecutor, as_completed

# cd "C:\Users\dawid.wiselka\OneDrive - NOMAD ELECTRIC Sp. z o.o\Dokumenty\Farmy\Updater\all\PLC-UPDATE"
# python FirmwareUpdater_listaExcel.py
# pyinstaller --onefile --noconsole --icon="plcv2.ico" --add-data "plcv2.ico;." --add-data "Default.scm.config;." FirmwareUpdater_listaExcel.py




# Konfiguracja hardcoded (nie zmienia się)
PLC_USER = "admin"
ROOT_PASS = "12345"
TIMEZONE = "Europe/Warsaw"
SYSTEM_SERVICES_FILE = "Default.scm.config"

# Domyślne wartości (będą w GUI)
DEFAULT_SSH_TIMEOUT = 30
DEFAULT_SSH_KEEPALIVE = 30
DEFAULT_RETRY_ATTEMPTS = 3
DEFAULT_RETRY_DELAY = 10
DEFAULT_PAUSE_BETWEEN = 5
DEFAULT_UPLOAD_TIMEOUT = 900  # 15 minut dla 300MB firmware
DEFAULT_UPDATE_COMMAND_TIMEOUT = 600  # 10 minut dla update-axcf
DEFAULT_IDLE_TIMEOUT = 60
DEFAULT_POST_REBOOT_WAIT = 60
DEFAULT_POST_REBOOT_TIMEOUT = 300
DEFAULT_POST_REBOOT_POLL = 5
DEFAULT_PARALLEL_WORKERS = 1

def resource_path(relative_path):
    """Zwraca absolutną ścieżkę do pliku, działa również w exe PyInstaller."""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def clean_ip_address(text):
    """
    Ekstraktuje adres IP z różnych formatów:
    - https://192.168.1.100/config → 192.168.1.100
    - 192.168.1.100:8080 → 192.168.1.100
    - [192.168.1.100] → 192.168.1.100
    """
    import re
    if not text:
        return ""
    
    # Znajdź wzorzec IP (4 liczby oddzielone kropkami)
    match = re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', text)
    if match:
        ip = match.group(1)
        # Walidacja zakresów (0-255)
        try:
            parts = ip.split('.')
            if all(0 <= int(p) <= 255 for p in parts):
                return ip
        except:
            pass
    
    return text.strip()  # Fallback - zwróć oczyszczony tekst

class FatalUpdateError(Exception):
    """Błąd krytyczny - operacja nie powinna być ponawiana (bez retry)."""
    pass

class PLCDevice:
    """Klasa reprezentująca jeden sterownik PLC."""
    def __init__(self, name, ip, password):
        self.name = name
        self.ip = ip
        self.password = password
        self.firmware_version = ""
        self.timezone = ""
        self.system_services_ok = ""
        self.last_check = ""
        self.last_update = ""
        self.status = "Oczekuje"
        self.error_log = ""
        self.plc_model = ""
        self.plc_time = ""
        self.time_sync_error = False

class BatchProcessorApp(tk.Tk):
    """Główna aplikacja do przetwarzania wsadowego sterowników PLC."""

    def __init__(self):
        super().__init__()
        self.title("PLC Batch Updater - Phoenix Contact")
        self.geometry("1300x850")
        self.configure(bg="#F0F4F8")
        try:
            self.iconbitmap(resource_path("plcv2.ico"))
        except:
            pass

        style = ttk.Style()
        try:
            style.theme_use('clam')
        except:
            pass

        # Nowoczesny schemat kolorów
        style.configure('TNotebook', 
                       background="#F0F4F8", 
                       borderwidth=0,
                       tabmargins=[2, 5, 2, 0])
        
        style.configure('TNotebook.Tab', 
                       font=('Segoe UI', 11, 'bold'),
                       padding=(20, 10),
                       borderwidth=0)
        
        style.map('TNotebook.Tab',
                  background=[('selected', '#FFFFFF'), ('!selected', '#CBD5E1')],
                  foreground=[('selected', '#1E40AF'), ('!selected', '#475569')],
                  expand=[('selected', [1, 1, 1, 0])])

        # Nowoczesna tabela
        style.configure('Modern.Treeview',
                        font=('Segoe UI', 10),
                        rowheight=32,
                        background='#FFFFFF',
                        fieldbackground='#FFFFFF',
                        borderwidth=0)
        
        style.configure('Modern.Treeview.Heading',
                        font=('Segoe UI', 10, 'bold'),
                        background='#E2E8F0',
                        foreground='#1E293B',
                        relief='flat',
                        borderwidth=1)
        
        style.map('Modern.Treeview',
                  background=[('selected', '#DBEAFE')],
                  foreground=[('selected', '#1E293B')])
        
        style.map('Modern.Treeview.Heading',
                  background=[('active', '#CBD5E1')])
        
        # Progress bar style
        style.configure('Custom.Horizontal.TProgressbar',
                       troughcolor='#E2E8F0',
                       background='#3B82F6',
                       bordercolor='#CBD5E1',
                       lightcolor='#60A5FA',
                       darkcolor='#2563EB',
                       thickness=20)

        # Zmienne stanu
        self.excel_path = tk.StringVar()
        self.firmware_path = tk.StringVar()
        self.devices = []
        self.processing = False
        self.log_queue = queue.Queue()
        self.upload_log_progress = {}
        self.show_errors_only = tk.BooleanVar(value=False)
        
        # Konfigurowalne ustawienia (domyślne wartości)
        self.ssh_timeout = DEFAULT_SSH_TIMEOUT
        self.ssh_keepalive = DEFAULT_SSH_KEEPALIVE
        self.retry_attempts = DEFAULT_RETRY_ATTEMPTS
        self.retry_delay = DEFAULT_RETRY_DELAY
        self.pause_between_devices = DEFAULT_PAUSE_BETWEEN
        self.upload_timeout = DEFAULT_UPLOAD_TIMEOUT
        self.update_command_timeout = DEFAULT_UPDATE_COMMAND_TIMEOUT
        self.idle_timeout = DEFAULT_IDLE_TIMEOUT
        self.post_reboot_wait = DEFAULT_POST_REBOOT_WAIT
        self.post_reboot_timeout = DEFAULT_POST_REBOOT_TIMEOUT
        self.post_reboot_poll = DEFAULT_POST_REBOOT_POLL
        self.parallel_workers = DEFAULT_PARALLEL_WORKERS
        
        # Tworzenie GUI
        self.create_widgets()

        self.firmware_path.trace_add("write", lambda *_: self.update_action_buttons_state())
        self.excel_path.trace_add("write", lambda *_: self.update_action_buttons_state())
        self.update_action_buttons_state()
        
        # Timer do aktualizacji logów
        self.update_logs()

    def create_action_button(self, parent, text, command, variant="neutral", **kwargs):
        """Tworzy nowoczesny przycisk z lepszym designem."""
        palette = {
            "neutral": {"bg": "#E2E8F0", "fg": "#1E293B", "active": "#CBD5E1", "border": "#94A3B8"},
            "primary": {"bg": "#3B82F6", "fg": "#FFFFFF", "active": "#2563EB", "border": "#1D4ED8"},
            "success": {"bg": "#10B981", "fg": "#FFFFFF", "active": "#059669", "border": "#047857"},
            "warning": {"bg": "#F59E0B", "fg": "#FFFFFF", "active": "#D97706", "border": "#B45309"},
            "danger": {"bg": "#EF4444", "fg": "#FFFFFF", "active": "#DC2626", "border": "#B91C1C"},
            "info": {"bg": "#06B6D4", "fg": "#FFFFFF", "active": "#0891B2", "border": "#0E7490"},
            "accent": {"bg": "#8B5CF6", "fg": "#FFFFFF", "active": "#7C3AED", "border": "#6D28D9"}
        }

        color = palette.get(variant, palette["neutral"])
        style = ttk.Style()
        style_name = f"Modern.{variant}.TButton"
        style.configure(
            style_name,
            font=("Segoe UI", 11, "bold"),
            padding=(16, 10),
            foreground=color["fg"],
            background=color["bg"],
            relief="flat",
            borderwidth=1,
            bordercolor=color["border"]
        )
        style.map(
            style_name,
            background=[('active', color["active"]), ('pressed', color["active"]), ('disabled', '#CBD5E1')],
            foreground=[('disabled', '#94A3B8')],
            relief=[('pressed', 'flat'), ('!pressed', 'flat')]
        )

        btn = ttk.Button(
            parent,
            text=text,
            command=command,
            style=style_name,
            cursor="hand2",
            **kwargs
        )
        return btn


    def create_ssh_client(self, ip, password, timeout=None):
        """Tworzy i konfiguruje klienta SSH z odpowiednimi timeoutami."""
        if timeout is None:
            timeout = self.ssh_timeout

        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(
                ip,
                username=PLC_USER,
                password=password,
                timeout=timeout,
                banner_timeout=timeout,
                auth_timeout=timeout,
                allow_agent=False,
                look_for_keys=False
            )

            transport = ssh.get_transport()
            if transport:
                transport.set_keepalive(self.ssh_keepalive)

            return ssh
        except paramiko.AuthenticationException as e:
            diagnosis = self.diagnose_ssh_error(ip, e, timeout)
            raise Exception(f"{diagnosis}: {str(e)}") from e
        except ConnectionRefusedError as e:
            diagnosis = self.diagnose_ssh_error(ip, e, timeout)
            raise Exception(f"{diagnosis}: {str(e)}") from e
        except socket.timeout as e:
            diagnosis = self.diagnose_ssh_error(ip, e, timeout)
            raise Exception(f"{diagnosis}: {str(e)}") from e
        except TimeoutError as e:
            diagnosis = self.diagnose_ssh_error(ip, e, timeout)
            raise Exception(f"{diagnosis}: {str(e)}") from e
        except Exception as e:
            diagnosis = self.diagnose_ssh_error(ip, e, timeout)
            raise Exception(f"{diagnosis}: {str(e)}") from e

    def check_ping(self, ip):
        """Sprawdza, czy host odpowiada na ping."""
        try:
            result = subprocess.run(
                ["ping", "-n", "1", "-w", "1200", ip],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False
            )
            return result.returncode == 0
        except Exception:
            return None

    def check_ssh_port(self, ip, timeout=None):
        """Sprawdza dostępność portu SSH 22."""
        if timeout is None:
            timeout = self.ssh_timeout

        sock = None
        try:
            sock = socket.create_connection((ip, 22), timeout=timeout)
            return True, None
        except ConnectionRefusedError:
            return False, "Port zamknięty (firewall)"
        except socket.timeout:
            return False, "Timeout połączenia"
        except OSError:
            return False, "Host nieosiągalny (ping fail)"
        finally:
            if sock:
                try:
                    sock.close()
                except Exception:
                    pass

    def diagnose_ssh_error(self, ip, error, timeout=None):
        """Diagnostyka błędów SSH z rozróżnieniem przyczyn."""
        if timeout is None:
            timeout = self.ssh_timeout

        error_msg = str(error).lower()

        if isinstance(error, paramiko.AuthenticationException) or "authentication failed" in error_msg:
            return "Błędne hasło"

        ping_ok = self.check_ping(ip)
        if ping_ok is False:
            return "Host nieosiągalny (ping fail)"

        port_open, port_reason = self.check_ssh_port(ip, timeout=min(timeout, 5))
        if not port_open:
            return port_reason

        if isinstance(error, socket.timeout) or "timed out" in error_msg or "timeout" in error_msg:
            return "Timeout połączenia"

        return "Błąd połączenia SSH"

    @contextmanager
    def ssh_connection(self, device):
        """
        Context manager dla bezpiecznego zarządzania połączeniem SSH.
        Automatycznie zamyka połączenie nawet przy błędach.
        
        Użycie:
            with self.ssh_connection(device) as (ssh, sftp):
                # ... operacje ...
        """
        ssh = None
        sftp = None
        
        try:
            self.log(f"  Otwieranie połączenia SSH do {device.ip}...")

            ssh = self.create_ssh_client(device.ip, device.password)
            
            sftp = ssh.open_sftp()
            
            self.log(f"  Połączono z {device.ip}")
            
            yield ssh, sftp
            
        except Exception as e:
            self.log(f"  Błąd połączenia SSH: {str(e)}")
            raise
            
        finally:
            # Zamknij SFTP
            if sftp:
                try:
                    sftp.close()
                    time.sleep(1)
                    self.log(f"  Zamknięto SFTP")
                    time.sleep(0.3)
                except Exception as e:
                    self.log(f"  UWAGA: Błąd zamykania SFTP: {str(e)}")
            
            # Zamknij SSH
            if ssh:
                try:
                    transport = ssh.get_transport()
                    if transport and transport.is_active():
                        transport.close()
                    ssh.close()
                    time.sleep(1)
                    self.log(f"  Zamknięto SSH")
                    time.sleep(1) 
                except Exception as e:
                    self.log(f"  UWAGA: Błąd zamykania SSH: {str(e)}")

    def wait_for_ssh_back(self, device):
        """Po restarcie czeka aktywnie na ponowną dostępność SSH sterownika."""
        max_attempts = max(1, int(self.post_reboot_timeout / self.post_reboot_poll))
        self.log(
            f"  Oczekiwanie po restarcie: start po {self.post_reboot_wait}s, "
            f"timeout globalny {self.post_reboot_timeout}s, "
            f"max prób reconnect: ~{max_attempts}"
        )
        time.sleep(self.post_reboot_wait)

        start_time = time.time()
        attempt = 0

        while (time.time() - start_time) < self.post_reboot_timeout:
            attempt += 1
            test_ssh = None
            try:
                test_ssh = paramiko.SSHClient()
                test_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                test_ssh.connect(
                    device.ip,
                    username=PLC_USER,
                    password=device.password,
                    timeout=10,
                    banner_timeout=10,
                    auth_timeout=10,
                    allow_agent=False,
                    look_for_keys=False
                )

                transport = test_ssh.get_transport()
                if transport:
                    transport.set_keepalive(self.ssh_keepalive)

                self.log(f"  Sterownik {device.ip} wrócił online (próba {attempt})")
                return True
            except (paramiko.AuthenticationException, ConnectionRefusedError, socket.timeout, TimeoutError, OSError) as e:
                elapsed = int(time.time() - start_time)
                reason = self.diagnose_ssh_error(device.ip, e, timeout=10)
                self.log(
                    f"  Reconnect próba {attempt}/{max_attempts} nieudana "
                    f"({elapsed}s/{self.post_reboot_timeout}s): {reason}"
                )
                time.sleep(self.post_reboot_poll)
            except Exception as e:
                elapsed = int(time.time() - start_time)
                self.log(
                    f"  Reconnect próba {attempt}/{max_attempts} nieudana "
                    f"({elapsed}s/{self.post_reboot_timeout}s): {str(e)}"
                )
                time.sleep(self.post_reboot_poll)
            finally:
                if test_ssh:
                    try:
                        test_ssh.close()
                    except:
                        pass

        raise Exception(
            f"Sterownik nie wrócił online po {self.post_reboot_timeout}s "
            f"od pierwszej próby połączenia (wykonano {attempt} prób reconnect)"
        )

    def is_transient_error(self, error):
        """Błędy tymczasowe - można ponawiać."""
        error_msg = str(error).lower()
        transient_keywords = [
            "timeout", "timed out", "eof", "socket", "connection reset",
            "connection refused", "network", "host unreachable", "banner"
        ]
        return any(keyword in error_msg for keyword in transient_keywords)

    def is_fatal_error(self, error):
        """Błędy krytyczne - bez retry."""
        return isinstance(error, FatalUpdateError)
        


    def execute_firmware_update(self, device):
        ssh = None
        channel = None
        try:
            device.status = "Aktualizacja firmware..."
            self.after(0, lambda d=device: self.update_device_row(d))
            self.log(f"  Nowe połączenie SSH dla firmware update...")

            ssh = self.create_ssh_client(device.ip, device.password)
            
            update_command = f"sudo update-axcf{device.plc_model}"
            self.log(f"  Uruchamiam: {update_command}")
            self.log(f"  Czekam na zakończenie procesu update (może zająć kilka minut)...")
            
            channel = ssh.get_transport().open_session()
            channel.get_pty()
            channel.exec_command(update_command)
            channel.send(device.password + "\n")
            
            output = ""
            start_time = time.time()
            timeout = self.update_command_timeout
            
            while True:
                if time.time() - start_time > timeout:
                    self.log(f"  UWAGA: Timeout - przekroczono {timeout}s oczekiwania")
                    break
                
                if channel.recv_ready():
                    chunk = channel.recv(1024).decode(errors="ignore")
                    output += chunk
                    for line in chunk.split('\n'):
                        if line.strip() and any(keyword in line.lower() for keyword in 
                            ['installing', 'updating', 'done', 'success', 'error', 'failed', 'reboot']):
                            self.log(f"    {line.strip()}")
                
                if channel.exit_status_ready():
                    exit_code = channel.recv_exit_status()
                    self.log(f"  Proces zakończony z kodem: {exit_code}")
                    
                    if exit_code != 0:
                            self.log(f"  UWAGA: Exit code: {exit_code} (może być normalne przy reboot)")
                    break
                
                time.sleep(0.5)
            
            if channel.recv_stderr_ready():
                errors = channel.recv_stderr(4096).decode(errors="ignore")
                if errors.strip():
                    self.log(f"  UWAGA: Stderr: {errors[:200]}")
            
            self.log("  Aktualizacja firmware zakończona. Sterownik restartuje się")
            device.status = "Oczekiwanie na restart..."
            self.after(0, lambda d=device: self.update_device_row(d))
            self.wait_for_ssh_back(device)
            
        except Exception as e:
            raise e
        finally:
            if channel:
                try:
                    channel.close()
                    self.log("  Zamknięto kanał SSH")
                except:
                    pass
            
            if ssh:
                try:
                    transport = ssh.get_transport()
                    if transport and transport.is_active():
                        transport.close()
                    ssh.close()
                    time.sleep(1)
                    self.log("  Zamknięto SSH")
                except:
                    pass
            
            time.sleep(3)

    def execute_reboot(self, device):
        ssh = None
        try:
            device.status = "Oczekiwanie na restart..."
            self.after(0, lambda d=device: self.update_device_row(d))
            self.log(f"  Nowe połączenie SSH dla reboot...")

            ssh = self.create_ssh_client(device.ip, device.password)
            
            self.log("  Uruchamiam 'sudo reboot'...")
            
            stdin, stdout, stderr = ssh.exec_command("sudo reboot", get_pty=True)
            stdin.write(device.password + "\n")
            stdin.flush()
            time.sleep(2)
            
        except Exception as e:
            # Ignoruj błędy zamknięcia - reboot ich powoduje
            if "Socket is closed" in str(e) or "Timeout" in str(e) or "EOF" in str(e):
                self.log("  Reboot zainicjowany (połączenie przerwane - oczekiwane)")
            else:
                raise e
        finally:
            if ssh:
                try:
                    ssh.close()
                    time.sleep(1)
                    self.log("  Zamknięto SSH po reboot")
                except:
                    pass
            time.sleep(1)

        self.wait_for_ssh_back(device)




    def upload_callback(self, filename, transferred, total, device=None):
        """
        Callback wywoływany podczas uploadu pliku przez SFTP.
        Aktualizuje progress bar i status.
        """
        if total > 0:
            percent = (transferred / total) * 100
            
            # Aktualizuj progress bar
            self.after(0, lambda p=percent: self.upload_progress.config(value=p))
            
            # Oblicz rozmiary w MB
            transferred_mb = transferred / 1024 / 1024
            total_mb = total / 1024 / 1024
            
            # Formatuj status
            status_text = f"{filename}: {transferred_mb:.1f} MB / {total_mb:.1f} MB ({percent:.1f}%)"

            if device:
                device.status = f"Wysyłanie pliku ({int(percent)}%)..."
                self.after(0, lambda d=device: self.update_device_row(d))
            
            # Aktualizuj GUI (thread-safe)
            self.after(0, lambda: self.upload_status_label.config(
                text=status_text, 
                fg="#3B82F6"
            ))
            
            # Log co 10%
            progress_threshold = int(percent // 10) * 10
            last_logged = self.upload_log_progress.get(filename, 0)

            if progress_threshold >= 10 and progress_threshold > last_logged:
                self.upload_log_progress[filename] = progress_threshold
                self.log(f"  Upload: {progress_threshold}% ({transferred_mb:.1f}/{total_mb:.1f} MB)")

    def upload_file_with_resume(self, sftp, local_path, remote_path, device=None):
        """Upload z obsługą .partial, wznowieniem i timeoutami postępu."""
        filename = os.path.basename(local_path)
        remote_partial_path = f"{remote_path}.partial"
        local_size = os.path.getsize(local_path)

        if local_size <= 0:
            raise Exception(f"Nieprawidłowy rozmiar pliku: {filename}")

        resume_offset = 0
        try:
            resume_offset = sftp.stat(remote_partial_path).st_size
        except FileNotFoundError:
            resume_offset = 0
        except IOError:
            resume_offset = 0

        if resume_offset > local_size:
            self.log(f"  UWAGA: Plik .partial większy niż lokalny. Usuwam i zaczynam od zera: {remote_partial_path}")
            sftp.remove(remote_partial_path)
            resume_offset = 0

        if resume_offset > 0:
            self.log(
                f"  Wznawianie transferu od {resume_offset/1024/1024:.1f} MB "
                f"z {local_size/1024/1024:.1f} MB"
            )
        else:
            self.log(f"  Start transferu: {filename} ({local_size/1024/1024:.1f} MB)")

        transfer_start = time.time()
        transferred = resume_offset
        chunk_size = 64 * 1024

        channel = sftp.get_channel()
        channel.settimeout(self.idle_timeout)

        mode = 'ab' if resume_offset > 0 else 'wb'
        try:
            with open(local_path, 'rb') as local_file:
                local_file.seek(resume_offset)
                with sftp.open(remote_partial_path, mode) as remote_file:
                    while transferred < local_size:
                        elapsed = time.time() - transfer_start
                        if elapsed > self.upload_timeout:
                            raise TimeoutError(
                                f"Timeout uploadu: przekroczono {self.upload_timeout}s "
                                f"dla pliku {filename}"
                            )

                        data = local_file.read(chunk_size)
                        if not data:
                            break

                        try:
                            remote_file.write(data)
                            remote_file.flush()
                        except socket.timeout as e:
                            raise TimeoutError(
                                f"Brak postępu transferu przez {self.idle_timeout}s "
                                f"(idle timeout)"
                            ) from e

                        transferred += len(data)
                        self.upload_callback(filename, transferred, local_size, device=device)
        finally:
            channel.settimeout(None)

        remote_partial_size = sftp.stat(remote_partial_path).st_size
        if remote_partial_size != local_size:
            raise Exception(
                f"Transfer niepełny! Lokalny: {local_size}, Zdalny .partial: {remote_partial_size}"
            )

        try:
            sftp.remove(remote_path)
        except FileNotFoundError:
            pass
        except IOError:
            pass

        sftp.rename(remote_partial_path, remote_path)

        remote_size = sftp.stat(remote_path).st_size
        if remote_size != local_size:
            raise Exception(f"Weryfikacja po rename nieudana! Lokalny: {local_size}, Zdalny: {remote_size}")

        self.log(f"  Transfer ukończony: {filename}")
        return remote_size

    def reset_upload_progress(self):
        """Resetuje progress bar po zakończeniu uploadu."""
        self.upload_log_progress.clear()
        self.after(0, lambda: self.upload_progress.config(value=0))
        self.after(0, lambda: self.upload_status_label.config(
            text="Oczekiwanie na transfer...",
            fg="#64748B"
        ))



    def process_batch(self, operation):
        """
        Główna metoda przetwarzania wsadowego.
        operation: "read", "system_services", "timezone", "firmware", "all"
        """
        self.processing = True
        self.after(0, self.update_action_buttons_state)
        
        total = len(self.devices)
        success_count = 0
        failed_count = 0
        failed_devices = []
        max_workers = max(1, min(5, int(self.parallel_workers)))

        def process_single_device(idx, device):
            if not self.processing:
                return "not_processed", "Operacja zatrzymana"

            self.log(f"\n{'='*60}")
            self.log(f"[{idx}/{total}] Przetwarzanie: {device.name} ({device.ip})")
            self.log(f"{'='*60}")

            device.status = "W trakcie"
            device.error_log = ""
            self.after(0, lambda d=device: self.update_device_row(d))

            attempt = 0
            success = False
            last_error = ""

            while attempt < self.retry_attempts and not success:
                if not self.processing:
                    return "not_processed", "Operacja zatrzymana"

                attempt += 1

                if attempt > 1:
                    self.log(
                        f"Retry próba {attempt}/{self.retry_attempts} "
                        f"(pozostało {self.retry_attempts - attempt + 1} prób)"
                    )
                    time.sleep(self.retry_delay)

                try:
                    if operation == "read":
                        self.read_single_device(device)
                        success = True

                    elif operation == "system_services":
                        self.update_system_services_only(device)
                        success = True

                    elif operation == "timezone":
                        self.update_timezone_only(device)
                        success = True

                    elif operation == "firmware":
                        self.update_firmware_only_operation(device)
                        success = True

                    elif operation == "all":
                        self.update_all_operations(device)
                        success = True

                    if success:
                        device.status = "OK"
                        self.log(f"[{device.name}] Operacja zakończona sukcesem")
                        return "success", ""

                except Exception as e:
                    error_msg = str(e)
                    last_error = error_msg
                    device.error_log = f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}: {error_msg}"

                    if self.is_fatal_error(e):
                        device.status = "Błąd"
                        self.log(f"[{device.name}] Błąd krytyczny (bez retry): {error_msg}")
                        return "failed", error_msg

                    if self.is_transient_error(e) and attempt < self.retry_attempts:
                        self.log(
                            f"Błąd tymczasowy (próba {attempt}/{self.retry_attempts}): {error_msg}"
                        )
                        self.log(
                            f"  Kolejny retry za {self.retry_delay}s "
                            f"(pozostało {self.retry_attempts - attempt} prób)"
                        )
                    else:
                        device.status = "Błąd"
                        if self.is_transient_error(e):
                            self.log(f"[{device.name}] Operacja nieudana po {self.retry_attempts} próbach: {error_msg}")
                        else:
                            self.log(f"[{device.name}] Błąd nienaprawialny (bez retry): {error_msg}")
                        return "failed", error_msg
                finally:
                    self.after(0, lambda d=device: self.update_device_row(d))

            if not success:
                device.status = "Błąd"
                return "failed", last_error or "Nieznany błąd"

            return "success", ""
        
        self.log(f"{'='*60}")
        self.log(f"START OPERACJI WSADOWEJ: {operation.upper()}")
        self.log(f"Liczba sterowników: {total}")
        self.log(f"Tryb równoległy: {max_workers} worker(ów)")
        self.log(f"{'='*60}")

        self.after(0, lambda: self.batch_progress.config(value=0))
        self.after(0, lambda: self.batch_progress_label.config(
            text=f"Start operacji {operation.upper()} (0/{total})",
            fg="#3B82F6"
        ))
        
        completed = 0
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {}

            for idx, device in enumerate(self.devices, 1):
                if not self.processing:
                    self.log("Operacja zatrzymana przez użytkownika")
                    break

                if idx > 1 and self.pause_between_devices > 0:
                    self.log(f"\nCzekam {self.pause_between_devices} sekund przed kolejnym sterownikiem...")
                    time.sleep(self.pause_between_devices)

                future = executor.submit(process_single_device, idx, device)
                futures[future] = device

            for future in as_completed(futures):
                device = futures[future]
                try:
                    result_status, error_msg = future.result()
                except Exception as e:
                    result_status, error_msg = "failed", str(e)

                if result_status == "success":
                    success_count += 1
                elif result_status == "failed":
                    failed_count += 1
                    failed_devices.append((device.name, error_msg))

                completed += 1
                progress_after = (completed / total) * 100 if total else 0
                self.after(0, lambda p=progress_after: self.batch_progress.config(value=p))
                self.after(0, lambda c=completed, t=total: self.batch_progress_label.config(
                    text=f"Postęp: {c}/{t} sterowników",
                    fg="#3B82F6"
                ))

        processed_count = success_count + failed_count
        not_processed_count = max(0, total - processed_count)

        recommendations = []
        if failed_devices:
            failed_text = "\n".join(msg for _, msg in failed_devices).lower()
            if "niezgodność" in failed_text or "kompatybil" in failed_text:
                recommendations.append("- Sprawdź zgodność modelu firmware (axcf2152/axcf3152).")
            if any(word in failed_text for word in ["timeout", "timed out", "connection", "socket", "eof"]):
                recommendations.append("- Sprawdź łączność sieciową i dostęp SSH do sterowników.")
            if "nie istnieje" in failed_text:
                recommendations.append("- Zweryfikuj obecność wymaganych plików lokalnych (firmware/System Services).")
        
        # Podsumowanie
        self.log(f"\n{'='*60}")
        self.log(f"PODSUMOWANIE OPERACJI: {operation.upper()}")
        self.log(f"{'='*60}")
        self.log(f"Sukces: {success_count}/{total}")
        self.log(f"Błędy: {failed_count}/{total}")
        self.log(f"Nieprzetworzone: {not_processed_count}/{total}")
        if failed_devices:
            self.log("Lista nieudanych sterowników:")
            for name, err in failed_devices[:10]:
                self.log(f"   - {name}: {err[:120]}")
            if len(failed_devices) > 10:
                self.log(f"   ... i {len(failed_devices) - 10} więcej")

        if recommendations:
            self.log("Rekomendacje:")
            for recommendation in recommendations:
                self.log(f"   {recommendation}")
        self.log(f"{'='*60}\n")
        
        self.processing = False
        self.after(0, self.update_action_buttons_state)
        self.after(0, lambda: self.status_bar.config(text="Gotowy"))
        self.after(0, lambda: self.batch_progress_label.config(
            text=f"Zakończono: sukces {success_count}, błędy {failed_count}, nieprzetworzone {not_processed_count}",
            fg="#10B981" if failed_count == 0 else "#EF4444"
        ))
        
        # Pokaż podsumowanie
        self.after(0, lambda: messagebox.showinfo(
            "Operacja zakończona",
            f"Operacja: {operation.upper()}\n\n"
            f"Sukces: {success_count}/{total}\n"
            f"Błędy: {failed_count}/{total}\n"
            f"Nieprzetworzone: {not_processed_count}/{total}\n\n"
            f"Sprawdź logi i zakładkę tabeli, aby uzyskać szczegóły."
        ))


    def read_single_device(self, device):
        """
        Odczytuje dane z pojedynczego sterownika.
        """
        try:
            device.status = "Łączenie SSH..."
            self.after(0, lambda d=device: self.update_device_row(d))
            
            with self.ssh_connection(device) as (ssh, sftp):
                
                # 1. Wykryj model PLC
                device.status = "Wykrywanie modelu..."
                self.after(0, lambda d=device: self.update_device_row(d))
                device.plc_model = self.detect_plc_model(ssh)
                
                # 2. Wersja Firmware
                device.status = "Odczyt firmware..."
                self.after(0, lambda d=device: self.update_device_row(d))
                stdin, stdout, stderr = ssh.exec_command("grep Arpversion /etc/plcnext/arpversion")
                fw_output = stdout.read().decode().strip()
                
                self.log(f"  Surowy output wersji firmware: '{fw_output}'")
                
                version_string = "?"
                if fw_output:
                    fw_output = fw_output.replace('Arpversion', '').strip()
                    
                    if ":" in fw_output:
                        parts = fw_output.split(':', 1)
                        version_string = parts[1].strip() if len(parts) > 1 else "?"
                    elif "=" in fw_output:
                        version_string = fw_output.split("=")[-1].strip()
                    else:
                        version_string = fw_output.strip()
                    
                    self.log(f"  Sparsowana wersja: '{version_string}'")
                
                if version_string and version_string != "?" and version_string[0].isdigit():
                    device.firmware_version = version_string
                else:
                    device.firmware_version = "?"
                    self.log(f"  UWAGA: Nie można odczytać poprawnej wersji firmware!")
                
                # 3. Strefa czasowa
                device.status = "Sprawdzanie strefy czasowej..."
                self.after(0, lambda d=device: self.update_device_row(d))
                stdin, stdout, stderr = ssh.exec_command("cat /etc/timezone")
                device.timezone = stdout.read().decode(errors="ignore").strip()
                
                # 4. Sprawdzenie synchronizacji czasu
                device.status = "Sprawdzanie synchronizacji czasu..."
                self.after(0, lambda d=device: self.update_device_row(d))
                plc_time_obj, plc_time_str, is_synced = self.check_time_sync(ssh)
                device.plc_time = plc_time_str
                device.time_sync_error = not is_synced
                
                # 5. System Services - porównanie zawartości pliku
                device.status = "Sprawdzanie System Services..."
                self.after(0, lambda d=device: self.update_device_row(d))
                try:
                    remote_path = "/opt/plcnext/config/System/Scm/Default.scm.config"
                    
                    local_file = resource_path(SYSTEM_SERVICES_FILE)
                    if os.path.exists(local_file):
                        # Odczytaj lokalny plik
                        with open(local_file, 'rb') as f:
                            local_content = f.read()
                        
                        # Pobierz zawartość zdalnego pliku
                        with sftp.open(remote_path, 'r') as remote_file:
                            remote_content = remote_file.read()
                        
                        # Porównaj zawartość (jako bytes)
                        if local_content == remote_content:
                            device.system_services_ok = "OK"
                            self.log(f"  System Services - zawartość zgodna")
                        else:
                            device.system_services_ok = "Niezgodność"
                            self.log(f"  UWAGA: System Services - zawartość różni się od wzorcowej")
                    else:
                        device.system_services_ok = "Brak lokalnego"
                        self.log(f"  UWAGA: Brak lokalnego pliku wzorcowego: {SYSTEM_SERVICES_FILE}")
                        
                except FileNotFoundError:
                    device.system_services_ok = "Brak"
                    self.log(f"  UWAGA: Plik System Services nie istnieje na sterowniku")
                except Exception as e:
                    device.system_services_ok = "Błąd"
                    self.log(f"  UWAGA: Błąd sprawdzania System Services: {str(e)}")
                
                # 6. Znacznik czasowy odczytu
                device.last_check = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # Logowanie podsumowania
                self.log(f"  Model: AXC F {device.plc_model}")
                self.log(f"  Firmware: {device.firmware_version}")
                self.log(f"  Czas PLC: {device.plc_time}")
                self.log(f"  Strefa czasowa: {device.timezone}")
                self.log(f"  System Services: {device.system_services_ok}")
                
            # Context manager automatycznie zamknie SSH/SFTP tutaj
            
        except Exception as e:
            raise e




    def detect_plc_model(self, ssh):
        """
        Wykrywa model sterownika PLC za pomocą komendy 'rauc status'.
        Zwraca numer modelu (np. "2152", "3152", "1152") lub None w przypadku błędu.
        """
        try:
            stdin, stdout, stderr = ssh.exec_command("rauc status")
            rauc_output = stdout.read().decode(errors="ignore").strip()
            
            # Szukamy linii "Compatible: axcfXXXX_v1"
            for line in rauc_output.split('\n'):
                if 'Compatible:' in line:
                    # Przykład: "Compatible: axcf2152_v1"
                    parts = line.split(':')
                    if len(parts) > 1:
                        compatible = parts[1].strip()
                        # Wyciągamy numer modelu (2152, 3152, 1152)
                        if 'axcf' in compatible:
                            model = compatible.replace('axcf', '').split('_')[0]
                            self.log(f"  Wykryty model PLC: AXC F {model}")
                            return model
            
            self.log(f"  UWAGA: Nie można wykryć modelu z 'rauc status'")
            return None
            
        except Exception as e:
            self.log(f"  UWAGA: Błąd wykrywania modelu: {str(e)}")
            return None

    def extract_model_from_firmware(self, firmware_path):
        """
        Wyciąga numer modelu z nazwy pliku firmware.
        Przykład: 'axcf2152-2024.0.8_LTS-24.0.8.183.raucb' -> '2152'
        """
        filename = os.path.basename(firmware_path)
        if filename.startswith('axcf'):
            model = filename.split('-')[0].replace('axcf', '')
            return model
        return None

    def validate_firmware_compatibility(self, device, firmware_path):
        """
        Sprawdza czy firmware jest kompatybilny z modelem sterownika.
        Zwraca (True, message) jeśli kompatybilny, (False, message) jeśli nie.
        """
        fw_model = self.extract_model_from_firmware(firmware_path)
        
        if not fw_model:
            return False, "Nie można odczytać modelu z nazwy firmware"
        
        if not device.plc_model:
            return False, "Model sterownika nie został wykryty"
        
        if fw_model != device.plc_model:
            return False, f"NIEZGODNOŚĆ: Firmware dla {fw_model}, sterownik to {device.plc_model}"
        
        return True, f"Firmware kompatybilny z modelem {device.plc_model}"

    def check_time_sync(self, ssh):
        """
        Sprawdza czy czas sterownika jest zsynchronizowany z czasem systemowym.
        Zwraca (datetime_object, time_string, is_synced).
        """
        try:
            # Pobierz czas z sterownika z timeoutem
            stdin, stdout, stderr = ssh.exec_command("date '+%Y-%m-%d %H:%M:%S'", timeout=10)
            plc_time_str = stdout.read().decode(errors="ignore").strip()
            
            if not plc_time_str:
                self.log(f"  UWAGA: Nie można odczytać czasu ze sterownika")
                return None, "", False
            
            # Parsuj czas sterownika
            plc_time = datetime.strptime(plc_time_str, "%Y-%m-%d %H:%M:%S")
            
            # Pobierz aktualny czas lokalny (warszawski)
            local_tz = pytz.timezone(TIMEZONE)
            local_time = datetime.now(local_tz).replace(tzinfo=None)
            
            # Oblicz różnicę
            time_diff = abs((local_time - plc_time).total_seconds())
            
            # Tolerancja 60 sekund
            is_synced = time_diff < 60
            
            if not is_synced:
                self.log(f"  UWAGA: DESYNCHRONIZACJA CZASU: różnica {time_diff:.0f}s")
                self.log(f"    Sterownik: {plc_time_str}")
                self.log(f"    Lokalny: {local_time.strftime('%Y-%m-%d %H:%M:%S')}")
            
            return plc_time, plc_time_str, is_synced
            
        except Exception as e:
            self.log(f"  UWAGA: Błąd sprawdzania czasu: {str(e)}")
            return None, "", False

    def compare_firmware_versions(self, current_version, target_version):
        """
        Porównuje wersje firmware ze szczegółowym logowaniem.
        Zwraca True jeśli wersje są IDENTYCZNE (nie trzeba aktualizować).
        """
        target_version_number = self.get_target_fw_version(target_version)
        
        self.log(f"  Porównanie wersji firmware:")
        self.log(f"     Aktualna wersja na sterowniku: '{current_version}'")
        self.log(f"     Wersja z pliku firmware: '{target_version_number}'")
        
        if not current_version or current_version == "?":
            self.log(f"     UWAGA: Nie można odczytać aktualnej wersji - wymuszam aktualizację")
            return False 
        
        if not target_version_number:
            self.log(f"     UWAGA: Nie można odczytać wersji z pliku - wymuszam aktualizację")
            return False
        
        # Normalizacja: usuń białe znaki i porównaj
        current_clean = current_version.strip()
        target_clean = target_version_number.strip()
        
        is_same = current_clean == target_clean
        
        if is_same:
            self.log(f"     Wersje są IDENTYCZNE - aktualizacja NIE jest potrzebna")
        else:
            self.log(f"     UWAGA: Wersje są RÓŻNE - aktualizacja jest potrzebna")
            self.log(f"        Różnica: '{current_clean}' != '{target_clean}'")
        
        return is_same
    
    def get_target_fw_version(self, firmware_path):
        """Wyodrębnia numer wersji z nazwy pliku firmware."""
        # Przykład: 'axcf2152-2024.0.8_LTS-24.0.8.183.raucb' -> '24.0.8.183'
        filename = os.path.basename(firmware_path)
        
        # Usuń rozszerzenie .raucb
        if filename.endswith('.raucb'):
            filename = filename[:-6]
        
        # Podziel po myślniku
        parts = filename.split('-')
        
        # Ostatnia część to wersja (np. '24.0.8.183')
        if len(parts) >= 3:
            version = parts[-1]
            self.log(f"  Wykryta wersja firmware z pliku: {version}")
            return version
        
        self.log(f"  UWAGA: Nie można odczytać wersji z nazwy pliku: {filename}")
        return ""

    def create_widgets(self):
        """Tworzy nowoczesny interfejs użytkownika."""
        
        # Notebook (zakładki)
        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=8, pady=8)
        
        # ZAKŁADKA 1: Przetwarzanie wsadowe
        batch_frame = tk.Frame(notebook, bg="#F8FAFC")
        notebook.add(batch_frame, text="Przetwarzanie wsadowe")
        
        # Sekcja pliku Excel
        excel_frame = tk.LabelFrame(batch_frame, 
                                    text="  Plik Excel z listą sterowników  ", 
                                    padx=15, pady=12,
                                    font=('Segoe UI', 11, 'bold'),
                                    fg="#1E293B",
                                    bg="#F8FAFC",
                                    relief="solid",
                                    borderwidth=1)
        excel_frame.pack(fill="x", padx=12, pady=8)
        
        self.create_action_button(
            excel_frame,
            text="Wybierz plik Excel",
            command=self.select_excel,
            variant="neutral"
        ).pack(side="left", padx=5)

        tk.Label(excel_frame, 
                textvariable=self.excel_path, 
                bg="#FFFFFF", 
                fg="#475569",
                relief="groove", 
                borderwidth=1,
                font=('Segoe UI', 10),
                width=55,
                anchor="w",
                padx=8, pady=6).pack(side="left", padx=8)

        self.load_excel_btn = self.create_action_button(
            excel_frame,
            text="Wczytaj listę",
            command=self.load_excel,
            variant="primary"
        )
        self.load_excel_btn.pack(side="left", padx=5)
        
        # Sekcja firmware
        firmware_frame = tk.LabelFrame(batch_frame, 
                                       text="  Plik Firmware (opcjonalnie dla aktualizacji)  ", 
                                       padx=15, pady=12,
                                       font=('Segoe UI', 11, 'bold'),
                                       fg="#1E293B",
                                       bg="#F8FAFC",
                                       relief="solid",
                                       borderwidth=1)
        firmware_frame.pack(fill="x", padx=12, pady=8)
        
        self.create_action_button(
            firmware_frame,
            text="Wybierz firmware",
            command=self.select_firmware,
            variant="neutral"
        ).pack(side="left", padx=5)

        tk.Label(firmware_frame, 
                textvariable=self.firmware_path, 
                bg="#FFFFFF", 
                fg="#475569",
                relief="groove", 
                borderwidth=1,
                font=('Segoe UI', 10),
                width=55,
                anchor="w",
                padx=8, pady=6).pack(side="left", padx=8)
        """
        # Typ sterownika
        plc_frame = tk.LabelFrame(batch_frame, text="Typ sterownika", padx=10, pady=5)
        plc_frame.pack(fill="x", padx=10, pady=5)
        tk.Radiobutton(plc_frame, text="AXC F 2152", variable=self.plc_type_var, value="2152").pack(side="left", padx=10)
        tk.Radiobutton(plc_frame, text="AXC F 3152", variable=self.plc_type_var, value="3152").pack(side="left", padx=10)
        """
        # Przyciski akcji - ODCZYT
        read_frame = tk.LabelFrame(batch_frame, 
                                  text="  Odczyt danych  ", 
                                  padx=15, pady=10,
                                  font=('Segoe UI', 11, 'bold'),
                                  fg="#1E293B",
                                  bg="#F8FAFC",
                                  relief="solid",
                                  borderwidth=1)
        read_frame.pack(fill="x", padx=12, pady=8)
        self.batch_read_btn = self.create_action_button(
            read_frame,
            text="Odczytaj wszystkie sterowniki",
            command=self.batch_read_all,
            variant="success"
        )
        self.batch_read_btn.pack(fill="x", padx=5, pady=5)

        # Przyciski akcji - AKTUALIZACJE (osobne)
        update_frame = tk.LabelFrame(batch_frame, 
                                    text="  Aktualizacje (wykonywane osobno)  ", 
                                    padx=15, pady=10,
                                    font=('Segoe UI', 11, 'bold'),
                                    fg="#1E293B",
                                    bg="#F8FAFC",
                                    relief="solid",
                                    borderwidth=1)
        update_frame.pack(fill="x", padx=12, pady=8)
        
        btn_grid = tk.Frame(update_frame, bg="#F8FAFC")
        btn_grid.pack(fill="x", padx=5, pady=5)
        
        self.batch_sys_btn = self.create_action_button(
            btn_grid,
            text="Wyślij System Services (wszystkie)",
            command=self.batch_system_services,
            variant="info"
        )
        self.batch_sys_btn.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        self.batch_tz_btn = self.create_action_button(
            btn_grid,
            text="Ustaw strefę czasową (wszystkie)",
            command=self.batch_timezone,
            variant="warning"
        )
        self.batch_tz_btn.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        self.batch_fw_btn = self.create_action_button(
            btn_grid,
            text="Aktualizuj Firmware (wszystkie)",
            command=self.batch_firmware_only,
            variant="primary"
        )
        self.batch_fw_btn.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

        self.batch_all_btn = self.create_action_button(
            btn_grid,
            text="WYKONAJ WSZYSTKO NARAZ",
            command=self.batch_update_all,
            variant="accent"
        )
        self.batch_all_btn.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        
        btn_grid.columnconfigure(0, weight=1)
        btn_grid.columnconfigure(1, weight=1)
        
        progress_frame = tk.LabelFrame(batch_frame, 
                                       text="  Status transferu plików  ", 
                                       padx=15, pady=10,
                                       font=('Segoe UI', 11, 'bold'),
                                       fg="#1E293B",
                                       bg="#F8FAFC",
                                       relief="solid",
                                       borderwidth=1)
        progress_frame.pack(fill="x", padx=12, pady=8)
        
        # Progress bar
        self.upload_progress = ttk.Progressbar(
            progress_frame, 
            orient="horizontal", 
            length=100, 
            mode="determinate",
            style="Custom.Horizontal.TProgressbar"
        )
        self.upload_progress.pack(fill="x", padx=5, pady=5)
        
        # Label ze statusem
        self.upload_status_label = tk.Label(
            progress_frame, 
            text="Oczekiwanie na transfer...",
            font=("Segoe UI", 10),
            fg="#64748B",
            bg="#F8FAFC"
        )
        self.upload_status_label.pack(padx=5, pady=4)

        batch_progress_frame = tk.LabelFrame(batch_frame, 
                                            text="  Postęp operacji wsadowej  ", 
                                            padx=15, pady=10,
                                            font=('Segoe UI', 11, 'bold'),
                                            fg="#1E293B",
                                            bg="#F8FAFC",
                                            relief="solid",
                                            borderwidth=1)
        batch_progress_frame.pack(fill="x", padx=12, pady=8)

        self.batch_progress = ttk.Progressbar(
            batch_progress_frame,
            orient="horizontal",
            length=100,
            mode="determinate",
            style="Custom.Horizontal.TProgressbar"
        )
        self.batch_progress.pack(fill="x", padx=5, pady=5)

        self.batch_progress_label = tk.Label(
            batch_progress_frame,
            text="Oczekiwanie na start...",
            font=("Segoe UI", 10),
            fg="#64748B",
            bg="#F8FAFC"
        )
        self.batch_progress_label.pack(padx=5, pady=4)
    

        control_frame = tk.Frame(batch_frame, bg="#F8FAFC")
        control_frame.pack(fill="x", padx=12, pady=8)

        self.save_excel_btn = self.create_action_button(
            control_frame,
            text="Zapisz raport Excel",
            command=self.save_excel,
            variant="primary"
        )
        self.save_excel_btn.pack(side="left", padx=5, fill="x", expand=True)

        self.stop_btn = self.create_action_button(
            control_frame,
            text="STOP",
            command=self.stop_processing,
            variant="danger",
            state="disabled"
        )
        self.stop_btn.pack(side="left", padx=5, fill="x", expand=True)

        filter_frame = tk.Frame(batch_frame, bg="#F8FAFC")
        filter_frame.pack(fill="x", padx=12, pady=(0, 5))
        tk.Checkbutton(
            filter_frame,
            text="Pokaż tylko sterowniki z problemami",
            variable=self.show_errors_only,
            command=self.refresh_device_tree,
            font=('Segoe UI', 10),
            bg="#F8FAFC",
            fg="#1E293B",
            selectcolor="#FFFFFF",
            activebackground="#F8FAFC",
            relief="flat"
        ).pack(side="left", padx=5)
        
        # Tabela ze sterownikami
        table_frame = tk.LabelFrame(batch_frame, 
                                   text="  Lista sterowników  ", 
                                   padx=10, pady=10,
                                   font=('Segoe UI', 11, 'bold'),
                                   fg="#1E293B",
                                   bg="#F8FAFC",
                                   relief="solid",
                                   borderwidth=1)
        table_frame.pack(fill="both", expand=True, padx=12, pady=8)

            # Scrollbar
        table_scroll_y = tk.Scrollbar(table_frame, orient="vertical")
        table_scroll_x = tk.Scrollbar(table_frame, orient="horizontal")

        self.device_tree = ttk.Treeview(table_frame, 
                                columns=("IP", "Model", "Firmware", "PLCTime", "Timezone", "SysServices", "LastCheck", "Status", "Issues"),
                                show="tree headings",
                                style='Modern.Treeview',
                                yscrollcommand=table_scroll_y.set,
                                xscrollcommand=table_scroll_x.set)

        table_scroll_y.config(command=self.device_tree.yview)
        table_scroll_x.config(command=self.device_tree.xview)

        self.device_tree.heading("#0", text="Nazwa")
        self.device_tree.heading("IP", text="IP")
        self.device_tree.heading("Model", text="Model PLC")
        self.device_tree.heading("Firmware", text="Wersja Firmware")
        self.device_tree.heading("PLCTime", text="Czas sterownika")
        self.device_tree.heading("Timezone", text="Strefa czasowa")
        self.device_tree.heading("SysServices", text="System Services")
        self.device_tree.heading("LastCheck", text="Ostatni odczyt")
        self.device_tree.heading("Status", text="Status")
        self.device_tree.heading("Issues", text="Issues") 

        self.device_tree.column("#0", width=150)
        self.device_tree.column("IP", width=90)
        self.device_tree.column("Model", width=80)
        self.device_tree.column("Firmware", width=100)
        self.device_tree.column("PLCTime", width=150)
        self.device_tree.column("Timezone", width=120)
        self.device_tree.column("SysServices", width=100)
        self.device_tree.column("LastCheck", width=150)
        self.device_tree.column("Status", width=80)
        self.device_tree.column("Issues", width=150)

        # Konfiguracja tagów dla kolorowania
        self.device_tree.tag_configure('success', background='#D1FAE5', foreground='#065F46')
        self.device_tree.tag_configure('error', background='#FEE2E2', foreground='#991B1B')
        self.device_tree.tag_configure('has_issues', background='#FEF3C7', foreground='#92400E')

        self.device_tree.pack(side="left", fill="both", expand=True)
        table_scroll_y.pack(side="right", fill="y")
        table_scroll_x.pack(side="bottom", fill="x")


        # ZAKŁADKA 2: Logi
        log_frame = tk.Frame(notebook, bg="#F8FAFC")
        notebook.add(log_frame, text="Logi operacji")

        self.log_text = scrolledtext.ScrolledText(log_frame, 
                                                   wrap=tk.WORD, 
                                                   font=("Consolas", 10),
                                                   bg="#1E293B",
                                                   fg="#E2E8F0",
                                                   insertbackground="#60A5FA",
                                                   relief="flat",
                                                   borderwidth=0)
        self.log_text.pack(fill="both", expand=True, padx=12, pady=12)

        self.create_action_button(
            log_frame,
            text="Wyczysc logi",
            command=self.clear_logs,
            variant="neutral"
        ).pack(pady=8)

        # ZAKŁADKA 3: Konfiguracja
        config_frame = tk.Frame(notebook, bg="#F8FAFC")
        notebook.add(config_frame, text="Konfiguracja")
        self.create_config_interface(config_frame)

        # ZAKŁADKA 4: Ręczna obsługa (poprawiona)
        manual_frame = tk.Frame(notebook, bg="#F8FAFC")
        notebook.add(manual_frame, text="Reczna obsluga")
        self.create_manual_interface(manual_frame)

        # Status bar
        self.status_bar = tk.Label(self, 
                                   text="Gotowy", 
                                   relief="flat", 
                                   anchor="w", 
                                   bg="#E2E8F0",
                                   fg="#1E293B",
                                   font=("Segoe UI", 10),
                                   padx=10, pady=6)
        self.status_bar.pack(side="bottom", fill="x")

    def update_action_buttons_state(self):
        """Włącza/wyłącza przyciski zgodnie z aktualnym etapem pracy."""
        has_devices = len(self.devices) > 0
        has_firmware = bool(self.firmware_path.get() and os.path.exists(self.firmware_path.get()))
        is_busy = self.processing

        normal = "normal"
        disabled = "disabled"

        if hasattr(self, 'load_excel_btn'):
            self.load_excel_btn.config(state=disabled if is_busy else normal)
        if hasattr(self, 'batch_read_btn'):
            self.batch_read_btn.config(state=normal if (has_devices and not is_busy) else disabled)
        if hasattr(self, 'batch_sys_btn'):
            self.batch_sys_btn.config(state=normal if (has_devices and not is_busy) else disabled)
        if hasattr(self, 'batch_tz_btn'):
            self.batch_tz_btn.config(state=normal if (has_devices and not is_busy) else disabled)
        if hasattr(self, 'batch_fw_btn'):
            self.batch_fw_btn.config(state=normal if (has_devices and has_firmware and not is_busy) else disabled)
        if hasattr(self, 'batch_all_btn'):
            self.batch_all_btn.config(state=normal if (has_devices and has_firmware and not is_busy) else disabled)
        if hasattr(self, 'save_excel_btn'):
            self.save_excel_btn.config(state=normal if (has_devices and not is_busy) else disabled)
        if hasattr(self, 'stop_btn'):
            self.stop_btn.config(state=normal if is_busy else disabled)

    def device_has_issues(self, device):
        """Czy urządzenie ma problemy prezentowane w kolumnie Issues."""
        if device.time_sync_error:
            return True
        if device.system_services_ok not in ["OK", ""]:
            return True
        if device.timezone and device.timezone.strip() != TIMEZONE.strip():
            return True
        if device.status == "Błąd":
            return True
        return False

    def get_device_row_render_data(self, device):
        """Przygotowuje wartości i tagi dla jednego wiersza tabeli."""
        issues = []
        has_issues = False

        plc_time_display = device.plc_time
        if device.time_sync_error:
            plc_time_display = f"BŁĄD: {device.plc_time}"
            issues.append("Desynchronizacja czasu")
            has_issues = True

        sys_services_display = device.system_services_ok
        if device.system_services_ok not in ["OK", ""]:
            sys_services_display = f"BŁĄD: {device.system_services_ok}"
            issues.append("System Services")
            has_issues = True

        timezone_display = device.timezone
        if device.timezone and device.timezone.strip() != TIMEZONE.strip():
            timezone_display = f"BŁĄD: {device.timezone}"
            issues.append(f"Strefa czasowa ({device.timezone} ≠ {TIMEZONE})")
            has_issues = True

        if device.status == "W trakcie":
            issues_text = "Sprawdzanie..."
        elif issues:
            issues_text = "\n".join(issues)
        else:
            issues_text = "Brak"

        values = (
            device.ip,
            f"AXC F {device.plc_model}" if device.plc_model else "",
            device.firmware_version,
            plc_time_display,
            timezone_display,
            sys_services_display,
            device.last_check,
            device.status,
            issues_text
        )

        if has_issues:
            tags = ('has_issues',)
        elif device.status == "OK":
            tags = ('success',)
        elif device.status == "Błąd":
            tags = ('error',)
        else:
            tags = ()

        return values, tags

    def refresh_device_tree(self):
        """Odświeża tabelę urządzeń z uwzględnieniem filtra."""
        self.device_tree.delete(*self.device_tree.get_children())

        show_only_errors = self.show_errors_only.get()
        for device in self.devices:
            if show_only_errors and not self.device_has_issues(device):
                continue

            values, tags = self.get_device_row_render_data(device)
            self.device_tree.insert("", "end", text=device.name, values=values, tags=tags)

    def create_config_interface(self, parent):
        """Tworzy interfejs konfiguracji z edytowalnymi parametrami."""
        
        # Kontener główny z scrollem
        canvas = tk.Canvas(parent, bg="#F8FAFC", highlightthickness=0)
        scrollbar = tk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#F8FAFC")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # SSH Settings
        ssh_frame = tk.LabelFrame(scrollable_frame,
                                 text="  SSH Settings  ",
                                 padx=20, pady=15,
                                 font=('Segoe UI', 11, 'bold'),
                                 fg="#1E293B",
                                 bg="#F8FAFC",
                                 relief="solid",
                                 borderwidth=1)
        ssh_frame.pack(fill="x", padx=15, pady=10)
        
        # SSH Timeout
        tk.Label(ssh_frame, text="Connection Timeout:", 
                font=('Segoe UI', 10), bg="#F8FAFC", fg="#475569").grid(row=0, column=0, sticky="w", pady=5)
        self.ssh_timeout_var = tk.IntVar(value=self.ssh_timeout)
        tk.Spinbox(ssh_frame, from_=10, to=120, textvariable=self.ssh_timeout_var,
                  width=10, font=('Segoe UI', 10)).grid(row=0, column=1, padx=10, pady=5)
        tk.Label(ssh_frame, text="s", font=('Segoe UI', 10), bg="#F8FAFC", fg="#64748B").grid(row=0, column=2, sticky="w")
        
        # SSH Keepalive
        tk.Label(ssh_frame, text="Keepalive Interval:", 
                font=('Segoe UI', 10), bg="#F8FAFC", fg="#475569").grid(row=1, column=0, sticky="w", pady=5)
        self.ssh_keepalive_var = tk.IntVar(value=self.ssh_keepalive)
        tk.Spinbox(ssh_frame, from_=10, to=120, textvariable=self.ssh_keepalive_var,
                  width=10, font=('Segoe UI', 10)).grid(row=1, column=1, padx=10, pady=5)
        tk.Label(ssh_frame, text="s", font=('Segoe UI', 10), bg="#F8FAFC", fg="#64748B").grid(row=1, column=2, sticky="w")
        
        # Retry Settings
        retry_frame = tk.LabelFrame(scrollable_frame,
                                   text="  Retry Settings  ",
                                   padx=20, pady=15,
                                   font=('Segoe UI', 11, 'bold'),
                                   fg="#1E293B",
                                   bg="#F8FAFC",
                                   relief="solid",
                                   borderwidth=1)
        retry_frame.pack(fill="x", padx=15, pady=10)
        
        # Retry Attempts
        tk.Label(retry_frame, text="Retry Attempts:", 
                font=('Segoe UI', 10), bg="#F8FAFC", fg="#475569").grid(row=0, column=0, sticky="w", pady=5)
        self.retry_attempts_var = tk.IntVar(value=self.retry_attempts)
        tk.Spinbox(retry_frame, from_=1, to=10, textvariable=self.retry_attempts_var,
                  width=10, font=('Segoe UI', 10)).grid(row=0, column=1, padx=10, pady=5)
        tk.Label(retry_frame, text="razy", font=('Segoe UI', 10), bg="#F8FAFC", fg="#64748B").grid(row=0, column=2, sticky="w")
        
        # Retry Delay
        tk.Label(retry_frame, text="Retry Delay:", 
                font=('Segoe UI', 10), bg="#F8FAFC", fg="#475569").grid(row=1, column=0, sticky="w", pady=5)
        self.retry_delay_var = tk.IntVar(value=self.retry_delay)
        tk.Spinbox(retry_frame, from_=5, to=60, textvariable=self.retry_delay_var,
                  width=10, font=('Segoe UI', 10)).grid(row=1, column=1, padx=10, pady=5)
        tk.Label(retry_frame, text="s", font=('Segoe UI', 10), bg="#F8FAFC", fg="#64748B").grid(row=1, column=2, sticky="w")
        
        # Transfer Settings
        transfer_frame = tk.LabelFrame(scrollable_frame,
                                      text="  Transfer Settings  ",
                                      padx=20, pady=15,
                                      font=('Segoe UI', 11, 'bold'),
                                      fg="#1E293B",
                                      bg="#F8FAFC",
                                      relief="solid",
                                      borderwidth=1)
        transfer_frame.pack(fill="x", padx=15, pady=10)
        
        # Pause Between Devices
        tk.Label(transfer_frame, text="Pause Between Devices:", 
                font=('Segoe UI', 10), bg="#F8FAFC", fg="#475569").grid(row=0, column=0, sticky="w", pady=5)
        self.pause_between_var = tk.IntVar(value=self.pause_between_devices)
        tk.Spinbox(transfer_frame, from_=0, to=30, textvariable=self.pause_between_var,
                  width=10, font=('Segoe UI', 10)).grid(row=0, column=1, padx=10, pady=5)
        tk.Label(transfer_frame, text="s", font=('Segoe UI', 10), bg="#F8FAFC", fg="#64748B").grid(row=0, column=2, sticky="w")
        
        # Upload Timeout
        tk.Label(transfer_frame, text="Upload Timeout (firmware):", 
                font=('Segoe UI', 10), bg="#F8FAFC", fg="#475569").grid(row=1, column=0, sticky="w", pady=5)
        self.upload_timeout_var = tk.IntVar(value=self.upload_timeout)
        tk.Spinbox(transfer_frame, from_=300, to=3600, increment=300, textvariable=self.upload_timeout_var,
                  width=10, font=('Segoe UI', 10)).grid(row=1, column=1, padx=10, pady=5)
        tk.Label(transfer_frame, text="s (15min)", font=('Segoe UI', 10), bg="#F8FAFC", fg="#64748B").grid(row=1, column=2, sticky="w")
        
        # Idle Timeout
        tk.Label(transfer_frame, text="Idle Timeout (no progress):", 
                font=('Segoe UI', 10), bg="#F8FAFC", fg="#475569").grid(row=2, column=0, sticky="w", pady=5)
        self.idle_timeout_var = tk.IntVar(value=self.idle_timeout)
        tk.Spinbox(transfer_frame, from_=30, to=300, textvariable=self.idle_timeout_var,
                  width=10, font=('Segoe UI', 10)).grid(row=2, column=1, padx=10, pady=5)
        tk.Label(transfer_frame, text="s", font=('Segoe UI', 10), bg="#F8FAFC", fg="#64748B").grid(row=2, column=2, sticky="w")

        # Update Command Timeout
        tk.Label(transfer_frame, text="Update Command Timeout (update-axcf):", 
            font=('Segoe UI', 10), bg="#F8FAFC", fg="#475569").grid(row=3, column=0, sticky="w", pady=5)
        self.update_command_timeout_var = tk.IntVar(value=self.update_command_timeout)
        tk.Spinbox(transfer_frame, from_=300, to=1800, increment=60, textvariable=self.update_command_timeout_var,
              width=10, font=('Segoe UI', 10)).grid(row=3, column=1, padx=10, pady=5)
        tk.Label(transfer_frame, text="s (10min)", font=('Segoe UI', 10), bg="#F8FAFC", fg="#64748B").grid(row=3, column=2, sticky="w")
        
        # Reboot Settings
        reboot_frame = tk.LabelFrame(scrollable_frame,
                                    text="  Reboot Settings  ",
                                    padx=20, pady=15,
                                    font=('Segoe UI', 11, 'bold'),
                                    fg="#1E293B",
                                    bg="#F8FAFC",
                                    relief="solid",
                                    borderwidth=1)
        reboot_frame.pack(fill="x", padx=15, pady=10)
        
        # Post Reboot Wait
        tk.Label(reboot_frame, text="Initial Wait After Reboot:", 
                font=('Segoe UI', 10), bg="#F8FAFC", fg="#475569").grid(row=0, column=0, sticky="w", pady=5)
        self.post_reboot_wait_var = tk.IntVar(value=self.post_reboot_wait)
        tk.Spinbox(reboot_frame, from_=30, to=180, textvariable=self.post_reboot_wait_var,
                  width=10, font=('Segoe UI', 10)).grid(row=0, column=1, padx=10, pady=5)
        tk.Label(reboot_frame, text="s", font=('Segoe UI', 10), bg="#F8FAFC", fg="#64748B").grid(row=0, column=2, sticky="w")
        
        # Post Reboot Timeout
        tk.Label(reboot_frame, text="Reconnect Global Timeout:", 
                font=('Segoe UI', 10), bg="#F8FAFC", fg="#475569").grid(row=1, column=0, sticky="w", pady=5)
        self.post_reboot_timeout_var = tk.IntVar(value=self.post_reboot_timeout)
        tk.Spinbox(reboot_frame, from_=120, to=600, textvariable=self.post_reboot_timeout_var,
                  width=10, font=('Segoe UI', 10)).grid(row=1, column=1, padx=10, pady=5)
        tk.Label(reboot_frame, text="s", font=('Segoe UI', 10), bg="#F8FAFC", fg="#64748B").grid(row=1, column=2, sticky="w")
        
        # Post Reboot Poll
        tk.Label(reboot_frame, text="Poll Interval:", 
                font=('Segoe UI', 10), bg="#F8FAFC", fg="#475569").grid(row=2, column=0, sticky="w", pady=5)
        self.post_reboot_poll_var = tk.IntVar(value=self.post_reboot_poll)
        tk.Spinbox(reboot_frame, from_=3, to=30, textvariable=self.post_reboot_poll_var,
                  width=10, font=('Segoe UI', 10)).grid(row=2, column=1, padx=10, pady=5)
        tk.Label(reboot_frame, text="s", font=('Segoe UI', 10), bg="#F8FAFC", fg="#64748B").grid(row=2, column=2, sticky="w")
        
        # Parallel Processing
        parallel_frame = tk.LabelFrame(scrollable_frame,
                          text="  Parallel Processing  ",
                                      padx=20, pady=15,
                                      font=('Segoe UI', 11, 'bold'),
                                      fg="#1E293B",
                                      bg="#F8FAFC",
                                      relief="solid",
                                      borderwidth=1)
        parallel_frame.pack(fill="x", padx=15, pady=10)
        
        tk.Label(parallel_frame, text="Parallel PLC workers:", 
                font=('Segoe UI', 10), bg="#F8FAFC", fg="#475569").grid(row=0, column=0, sticky="w", pady=5)
        self.parallel_workers_var = tk.IntVar(value=self.parallel_workers)
        tk.Spinbox(parallel_frame, from_=1, to=5, textvariable=self.parallel_workers_var,
              width=10, font=('Segoe UI', 10)).grid(row=0, column=1, padx=10, pady=5, sticky="w")
        tk.Label(parallel_frame, text="(1-5)", 
            font=('Segoe UI', 10), bg="#F8FAFC", fg="#64748B").grid(row=0, column=2, sticky="w")
        
        # Przyciski
        button_frame = tk.Frame(scrollable_frame, bg="#F8FAFC")
        button_frame.pack(fill="x", padx=15, pady=20)
        
        self.create_action_button(
            button_frame,
            text="Zastosuj zmiany",
            command=self.apply_config,
            variant="primary"
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        self.create_action_button(
            button_frame,
            text="Przywroc domyslne",
            command=self.reset_config,
            variant="neutral"
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")

    def apply_config(self):
        """Zastosuj zmiany z zakładki konfiguracji."""
        self.ssh_timeout = self.ssh_timeout_var.get()
        self.ssh_keepalive = self.ssh_keepalive_var.get()
        self.retry_attempts = self.retry_attempts_var.get()
        self.retry_delay = self.retry_delay_var.get()
        self.pause_between_devices = self.pause_between_var.get()
        self.upload_timeout = self.upload_timeout_var.get()
        self.update_command_timeout = self.update_command_timeout_var.get()
        self.idle_timeout = self.idle_timeout_var.get()
        self.post_reboot_wait = self.post_reboot_wait_var.get()
        self.post_reboot_timeout = self.post_reboot_timeout_var.get()
        self.post_reboot_poll = self.post_reboot_poll_var.get()
        self.parallel_workers = self.parallel_workers_var.get()
        
        self.log("Zastosowano nowe ustawienia konfiguracji")
        messagebox.showinfo("Sukces", "Ustawienia zostaly zaktualizowane")

    def reset_config(self):
        """Przywróć domyślne wartości konfiguracji."""
        self.ssh_timeout_var.set(DEFAULT_SSH_TIMEOUT)
        self.ssh_keepalive_var.set(DEFAULT_SSH_KEEPALIVE)
        self.retry_attempts_var.set(DEFAULT_RETRY_ATTEMPTS)
        self.retry_delay_var.set(DEFAULT_RETRY_DELAY)
        self.pause_between_var.set(DEFAULT_PAUSE_BETWEEN)
        self.upload_timeout_var.set(DEFAULT_UPLOAD_TIMEOUT)
        self.update_command_timeout_var.set(DEFAULT_UPDATE_COMMAND_TIMEOUT)
        self.idle_timeout_var.set(DEFAULT_IDLE_TIMEOUT)
        self.post_reboot_wait_var.set(DEFAULT_POST_REBOOT_WAIT)
        self.post_reboot_timeout_var.set(DEFAULT_POST_REBOOT_TIMEOUT)
        self.post_reboot_poll_var.set(DEFAULT_POST_REBOOT_POLL)
        self.parallel_workers_var.set(DEFAULT_PARALLEL_WORKERS)
        
        self.apply_config()
        self.log("Przywrocono domyslne ustawienia")

    def _clean_ip_field(self, entry_widget):
        """Czyści pole IP z niepotrzebnych znaków."""
        current_value = entry_widget.get()
        cleaned = clean_ip_address(current_value)
        if cleaned != current_value:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, cleaned)
            self.log(f"Oczyszczono IP: '{current_value}' -> '{cleaned}'")

    def create_manual_interface(self, parent):
        """Tworzy nowoczesny interfejs do ręcznej obsługi pojedynczego sterownika."""
        
        connection_frame = tk.LabelFrame(parent, 
                                        text="  Połączenie  ", 
                                        padx=15, pady=15,
                                        font=('Segoe UI', 11, 'bold'),
                                        fg="#1E293B",
                                        bg="#F8FAFC",
                                        relief="solid",
                                        borderwidth=1)
        connection_frame.pack(fill="x", padx=12, pady=8)
        
        tk.Label(connection_frame, 
                text="Adres IP:", 
                font=('Segoe UI', 10, 'bold'),
                bg="#F8FAFC",
                fg="#1E293B").pack(pady=(0,5))
        self.ip_entry = tk.Entry(connection_frame, 
                                width=25,
                                font=('Segoe UI', 11),
                                relief="solid",
                                borderwidth=1)
        self.ip_entry.pack(pady=(0,10))
        
        # Auto-czyszczenie IP przy wklejaniu (Ctrl+V lub paste)
        def on_ip_paste(event):
            # Dodaj opóźnienie aby pozwolić na wklejenie
            self.after(50, lambda: self._clean_ip_field(self.ip_entry))
        
        self.ip_entry.bind('<Control-v>', on_ip_paste)
        self.ip_entry.bind('<Button-3>', on_ip_paste)  # prawy klik -> paste
        self.ip_entry.bind('<FocusOut>', lambda e: self._clean_ip_field(self.ip_entry))
        
        tk.Label(connection_frame, 
                text="Hasło:", 
                font=('Segoe UI', 10, 'bold'),
                bg="#F8FAFC",
                fg="#1E293B").pack(pady=(0,5))
        self.password_entry = tk.Entry(connection_frame, 
                                      show="*", 
                                      width=25,
                                      font=('Segoe UI', 11),
                                      relief="solid",
                                      borderwidth=1)
        self.password_entry.pack(pady=(0,10))
        
        # DODANE: Typ sterownika dla ręcznej obsługi
        tk.Label(connection_frame, 
                text="Typ sterownika:", 
                font=('Segoe UI', 10, 'bold'),
                bg="#F8FAFC",
                fg="#1E293B").pack(pady=(10, 5))
        self.manual_plc_type_var = tk.StringVar(value="2152")
        plc_manual_frame = tk.Frame(connection_frame, bg="#F8FAFC")
        plc_manual_frame.pack(pady=(0,10))
        tk.Radiobutton(plc_manual_frame, 
                      text="AXC F 2152", 
                      variable=self.manual_plc_type_var, 
                      value="2152",
                      font=('Segoe UI', 10),
                      bg="#F8FAFC",
                      fg="#1E293B",
                      selectcolor="#FFFFFF",
                      activebackground="#F8FAFC").pack(side="left", padx=10)
        tk.Radiobutton(plc_manual_frame, 
                      text="AXC F 3152", 
                      variable=self.manual_plc_type_var, 
                      value="3152",
                      font=('Segoe UI', 10),
                      bg="#F8FAFC",
                      fg="#1E293B",
                      selectcolor="#FFFFFF",
                      activebackground="#F8FAFC").pack(side="left", padx=10)
        
        self.create_action_button(
            connection_frame,
            text="Odczytaj dane z PLC",
            command=self.manual_read_plc,
            variant="primary"
        ).pack(pady=10)
        
        self.manual_data_label = tk.Label(parent, 
                                         text="Tutaj pojawią się dane z PLC.",
                                         bg="#FFFFFF",
                                         fg="#475569",
                                         relief="solid",
                                         borderwidth=1,
                                         justify="left",
                                         font=("Segoe UI", 10), 
                                         wraplength=500, 
                                         padx=15, pady=15,
                                         anchor="nw")
        self.manual_data_label.pack(fill="x", padx=12, pady=8)
        
        # Sekcja operacji ręcznych
        operations_frame = tk.LabelFrame(parent, 
                                        text="  Operacje pojedyncze  ", 
                                        padx=15, pady=12,
                                        font=('Segoe UI', 11, 'bold'),
                                        fg="#1E293B",
                                        bg="#F8FAFC",
                                        relief="solid",
                                        borderwidth=1)
        operations_frame.pack(fill="x", padx=12, pady=8)
        
        # Strefa czasowa
        self.create_action_button(
            operations_frame,
            text="Ustaw strefę czasową",
            command=self.manual_set_timezone,
            variant="warning"
        ).pack(fill="x", padx=5, pady=3)
        
        # System Services
        self.create_action_button(
            operations_frame,
            text="Wyślij System Services",
            command=self.manual_upload_system_services,
            variant="info"
        ).pack(fill="x", padx=5, pady=3)
        
        # Firmware
        firmware_manual_frame = tk.LabelFrame(parent, 
                                             text="  Aktualizacja Firmware  ", 
                                             padx=15, pady=12,
                                             font=('Segoe UI', 11, 'bold'),
                                             fg="#1E293B",
                                             bg="#F8FAFC",
                                             relief="solid",
                                             borderwidth=1)
        firmware_manual_frame.pack(fill="x", padx=12, pady=8)
        
        self.create_action_button(
            firmware_manual_frame,
            text="Wybierz plik firmware",
            command=self.select_manual_firmware,
            variant="neutral"
        ).pack(pady=8)
        self.manual_firmware_path = tk.StringVar()
        tk.Label(firmware_manual_frame, 
                textvariable=self.manual_firmware_path, 
                bg="#FFFFFF",
                fg="#475569",
                relief="solid",
                borderwidth=1,
                font=('Segoe UI', 10),
                wraplength=500,
                padx=10, pady=8,
                anchor="w").pack(pady=8, fill="x", padx=5)
        
        manual_fw_buttons = tk.Frame(firmware_manual_frame)
        manual_fw_buttons.pack(pady=5)
        self.create_action_button(
            manual_fw_buttons,
            text="Wyślij firmware",
            command=self.manual_upload_firmware,
            variant="success"
        ).pack(side="left", padx=5)
        self.create_action_button(
            manual_fw_buttons,
            text="Wykonaj aktualizację",
            command=self.manual_execute_update,
            variant="danger"
        ).pack(side="left", padx=5)

    def select_excel(self):
        """Wybór pliku Excel."""
        filepath = filedialog.askopenfilename(
            title="Wybierz plik Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filepath:
            self.excel_path.set(filepath)

    def select_firmware(self):
        """Wybór pliku firmware."""
        filepath = filedialog.askopenfilename(title="Wybierz plik firmware")
        if filepath:
            self.firmware_path.set(filepath)

    def load_excel(self):
        """Wczytuje listę sterowników z pliku Excel."""
        excel_file = self.excel_path.get()
        if not excel_file or not os.path.exists(excel_file):
            messagebox.showerror("Błąd", "Wybierz prawidłowy plik Excel!")
            return
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            self.devices = []
            
            # Pomijamy nagłówek (wiersz 1)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # Nazwa i IP muszą być wypełnione
                    name = str(row[0]).strip()
                    ip = str(row[1]).strip()
                    password = str(row[2]).strip() if row[2] else ""
                    
                    device = PLCDevice(name, ip, password)
                    
                    # Wczytaj istniejące dane jeśli są
                    if len(row) > 3 and row[3]:
                        device.firmware_version = str(row[3])
                    if len(row) > 4 and row[4]:
                        device.timezone = str(row[4])
                    if len(row) > 5 and row[5]:
                        device.system_services_ok = str(row[5])
                    if len(row) > 6 and row[6]:
                        device.last_check = str(row[6])
                    
                    self.devices.append(device)
            
            wb.close()
            self.refresh_device_tree()
            self.update_action_buttons_state()
            self.log(f"Wczytano {len(self.devices)} sterowników z pliku Excel")
            messagebox.showinfo("Sukces", f"Wczytano {len(self.devices)} sterowników")
            
        except Exception as e:
            self.log(f"Błąd wczytywania Excel: {str(e)}")
            messagebox.showerror("Błąd", f"Błąd wczytywania pliku Excel:\n{str(e)}")

    def save_excel(self):
        """Zapisuje aktualny stan do pliku Excel."""
        if not self.devices:
            messagebox.showwarning("Uwaga", "Brak danych do zapisania!")
            return
        
        try:
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"PLC_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not save_path:
                return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sterowniki PLC"
            
            # Nagłówki
            headers = ["Nazwa Farmy", "IP", "Hasło", "Firmware", "Strefa czasowa", 
                      "System Services", "Ostatni odczyt", "Ostatnia aktualizacja", "Status", "Logi błędów"]
            ws.append(headers)
            
            # Formatowanie nagłówków
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Dane
            for device in self.devices:
                ws.append([
                    device.name,
                    device.ip,
                    device.password,
                    device.firmware_version,
                    device.timezone,
                    device.system_services_ok,
                    device.last_check,
                    device.last_update,
                    device.status,
                    device.error_log
                ])
            
            # Dopasowanie szerokości kolumn
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(save_path)
            self.log(f"Zapisano raport do: {save_path}")
            messagebox.showinfo("Sukces", f"Raport zapisany:\n{save_path}")
            
        except Exception as e:
            self.log(f"Błąd zapisu Excel: {str(e)}")
            messagebox.showerror("Błąd", f"Błąd zapisu do Excel:\n{str(e)}")

    def update_firmware_only_operation(self, device):
        """
        Aktualizuje TYLKO firmware (z automatycznym wykrywaniem modelu i walidacją).
        POPRAWIONA: Używa execute_firmware_update() dla bezpiecznego reebootu.
        """
        self.log(f"Aktualizacja Firmware...")
        
        firmware_file = self.firmware_path.get()
        
        # Odczyt danych (w tym model PLC) - PRZERWIJ jeśli błąd
        try:
            self.read_single_device(device)
        except Exception as e:
            error_msg = f"Nie można odczytać danych sterownika przed aktualizacją: {str(e)}"
            self.log(f"  BŁĄD: {error_msg}")
            raise Exception(error_msg)
        
        # Walidacja kompatybilności
        is_compatible, compat_msg = self.validate_firmware_compatibility(device, firmware_file)
        self.log(f"  {compat_msg}")
        
        if not is_compatible:
            raise FatalUpdateError(compat_msg)
        
        # Sprawdź czy firmware jest aktualny
        if self.compare_firmware_versions(device.firmware_version, firmware_file):
            self.log(f"  Firmware już aktualny (v.{device.firmware_version}) - pomijam aktualizację")
            device.last_update = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            return True
        
        try:
            # KROK 1: UPLOAD FIRMWARE (w context managerze)
            with self.ssh_connection(device) as (ssh, sftp):
                
                filename = os.path.basename(firmware_file)
                remote_fw_path = f"/opt/plcnext/{filename}"
                
                file_size = os.path.getsize(firmware_file)
                self.log(f"  Wysyłanie firmware ({file_size/1024/1024:.1f} MB)...")
                
                self.upload_file_with_resume(
                    sftp,
                    firmware_file,
                    remote_fw_path,
                    device=device
                )
                
                self.reset_upload_progress()
                
                self.log(f"  Firmware wysłany i zweryfikowany")
            
            # Context manager zamknął SSH/SFTP tutaj
            
            # KROK 2: WYKONAJ UPDATE (NOWE połączenie SSH)
            self.execute_firmware_update(device)
            
            device.last_update = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            return True
            
        except Exception as e:
            self.reset_upload_progress()
            raise e


    def batch_read_all(self):
            """Odczytuje dane ze wszystkich sterowników."""
            if not self.devices:
                messagebox.showwarning("Uwaga", "Najpierw wczytaj listę sterowników!")
                return
            
            if self.processing:
                messagebox.showwarning("Uwaga", "Operacja już w toku!")
                return
            
            threading.Thread(target=self.process_batch, args=("read",), daemon=True).start()

    def batch_system_services(self):
        """Wysyła System Services do wszystkich sterowników."""
        if not self.devices:
            messagebox.showwarning("Uwaga", "Najpierw wczytaj listę sterowników!")
            return
        
        if self.processing:
            messagebox.showwarning("Uwaga", "Operacja już w toku!")
            return
        
        response = messagebox.askyesno(
            "Potwierdzenie",
            f"Czy wysłać System Services do {len(self.devices)} sterowników?\n\n"
            "Każdy sterownik zostanie zrestartowany po aktualizacji."
        )
        
        if response:
            threading.Thread(target=self.process_batch, args=("system_services",), daemon=True).start()

    def batch_timezone(self):
        """Ustawia strefę czasową na wszystkich sterownikach."""
        if not self.devices:
            messagebox.showwarning("Uwaga", "Najpierw wczytaj listę sterowników!")
            return
        
        if self.processing:
            messagebox.showwarning("Uwaga", "Operacja już w toku!")
            return
        
        response = messagebox.askyesno(
            "Potwierdzenie",
            f"Czy ustawić strefę czasową {TIMEZONE} na {len(self.devices)} sterownikach?\n\n"
            "Każdy sterownik zostanie zrestartowany."
        )
        
        if response:
            threading.Thread(target=self.process_batch, args=("timezone",), daemon=True).start()

    def batch_firmware_only(self):
        """Aktualizuje firmware na wszystkich sterownikach."""
        if not self.devices:
            messagebox.showwarning("Uwaga", "Najpierw wczytaj listę sterowników!")
            return
        
        if self.processing:
            messagebox.showwarning("Uwaga", "Operacja już w toku!")
            return
        
        firmware_file = self.firmware_path.get()
        if not firmware_file or not os.path.exists(firmware_file):
            messagebox.showerror("Błąd", "Wybierz prawidłowy plik firmware!")
            return
        
        response = messagebox.askyesno(
            "Potwierdzenie",
            f"Czy zaktualizować firmware na {len(self.devices)} sterownikach?\n\n"
            "Każdy sterownik zostanie zrestartowany po aktualizacji.\n"
            "To może zająć dużo czasu!"
        )
        
        if response:
            threading.Thread(target=self.process_batch, args=("firmware",), daemon=True).start()

    def batch_update_all(self):
        """WYKONUJE WSZYSTKIE OPERACJE NARAZ - zoptymalizowane."""
        if not self.devices:
            messagebox.showwarning("Uwaga", "Najpierw wczytaj listę sterowników!")
            return
        
        if self.processing:
            messagebox.showwarning("Uwaga", "Operacja już w toku!")
            return
        
        firmware_file = self.firmware_path.get()
        if not firmware_file or not os.path.exists(firmware_file):
            messagebox.showerror("Błąd", "Wybierz prawidłowy plik firmware!")
            return
        
        response = messagebox.askyesno(
            "Potwierdzenie",
            f"PEŁNA AKTUALIZACJA {len(self.devices)} sterowników:\n\n"
            "Operacje wykonywane dla każdego sterownika:\n"
            "1. System Services (jeśli potrzebne)\n"
            "2. Firmware - wysłanie i sudo update\n"
            "3. Strefa czasowa (jeśli potrzebne)\n"
            "4. Restart sterownika\n\n"
            "Operacja może zająć bardzo dużo czasu!\n\n"
            "Kontynuować?"
        )
        
        if response:
            threading.Thread(target=self.process_batch, args=("all",), daemon=True).start()


    def update_system_services_only(self, device):
        """
        Wysyła System Services i restartuje sterownik. Pomija, jeśli jest już OK.
        POPRAWIONA: Używa execute_reboot() dla bezpiecznego reebootu.
        """
        self.log(f"Aktualizacja System Services...")
        
        # Sprawdzenie statusu przed operacją
        try:
            self.read_single_device(device)
        except Exception as e:
            self.log(f"  UWAGA: Błąd odczytu przed aktualizacją SysServices: {str(e)}")
        
        # Logika pominięcia
        if device.system_services_ok == "OK":
            self.log(f"  INFO: System Services już aktualne - pomijam")
            device.last_update = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            return True
        
        try:
            # KROK 1: UPLOAD SYSTEM SERVICES (w context managerze)
            with self.ssh_connection(device) as (ssh, sftp):
                
                local_sys_file = resource_path(SYSTEM_SERVICES_FILE)
                if not os.path.exists(local_sys_file):
                    raise FatalUpdateError(f"Plik {SYSTEM_SERVICES_FILE} nie istnieje!")
                
                remote_sys_path = "/opt/plcnext/config/System/Scm/Default.scm.config"
                filename = os.path.basename(local_sys_file)
                file_size = os.path.getsize(local_sys_file)
                
                self.log(f"  Wysyłanie {filename} ({file_size/1024:.1f} KB)...")
                
                self.upload_file_with_resume(
                    sftp,
                    local_sys_file,
                    remote_sys_path,
                    device=device
                )
                
                self.reset_upload_progress()
                
                device.system_services_ok = "OK"
                self.log(f"  System Services wysłane i zweryfikowane")
            
            # Context manager zamknął SSH/SFTP tutaj
            
            # KROK 2: REBOOT (NOWE połączenie SSH)
            self.execute_reboot(device)
            
            device.last_update = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            return True
            
        except Exception as e:
            self.reset_upload_progress()
            raise e


    def update_timezone_only(self, device):
        """
        Ustawia strefę czasową i restartuje. Pomija, jeśli już OK.
        POPRAWIONA: Używa execute_reboot() dla bezpiecznego reebootu.
        """
        self.log(f"Aktualizacja strefy czasowej na {TIMEZONE}...")
        
        # Sprawdzenie statusu przed operacją
        try:
            self.read_single_device(device)
        except Exception as e:
            self.log(f"  UWAGA: Błąd odczytu przed aktualizacją Timezone: {str(e)}")
        
        # Logika pominięcia
        if device.timezone.strip() == TIMEZONE.strip():
            self.log(f"  INFO: Strefa czasowa już ustawiona na {TIMEZONE} - pomijam")
            device.last_update = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            return True
        
        try:
            # KROK 1: USTAWIENIE TIMEZONE (w context managerze)
            with self.ssh_connection(device) as (ssh, sftp):
                
                self.log(f"  Ustawianie strefy czasowej na {TIMEZONE}...")
                
                # Wpisanie TIMEZONE do /etc/timezone
                stdin, stdout, stderr = ssh.exec_command(
                    f"sudo sh -c 'echo {TIMEZONE} > /etc/timezone'", 
                    get_pty=True
                )
                stdin.write(device.password + "\n")
                stdin.flush()
                time.sleep(1)
                
                # Użycie timedatectl
                stdin, stdout, stderr = ssh.exec_command(
                    f"sudo timedatectl set-timezone {TIMEZONE}", 
                    get_pty=True
                )
                stdin.write(device.password + "\n")
                stdin.flush()
                time.sleep(1)
                
                device.timezone = TIMEZONE
                self.log("  Strefa czasowa ustawiona")
            
            # Context manager zamknął SSH/SFTP tutaj
            
            # KROK 2: REBOOT (NOWE połączenie SSH)
            self.execute_reboot(device)
            
            device.last_update = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            return True
            
        except Exception as e:
            raise e

    def update_all_operations(self, device):
        """
        Wykonuje wszystkie operacje: System Services, Firmware, Timezone.
        Zoptymalizowane pod kątem restartów i pomijania.
        
        KOLEJNOŚĆ OPERACJI:
        1. Połączenie SSH (przez context manager)
        2. Wykrycie modelu PLC i walidacja kompatybilności firmware
        3. Odczyt wstępny (stan SysServices, Timezone, Firmware)
        4. Aktualizacja System Services (tylko jeśli jest różnica/brak)
        5. Aktualizacja Firmware (tylko wysłanie pliku - jeśli konieczne i kompatybilne)
        6. Ustawienie strefy czasowej (tylko jeśli konieczne)
        7. ZAMKNIĘCIE SFTP przed rebootem/update
        8. Wykonanie sudo update / sudo reboot (tylko jeśli FW lub SS było wgrywane)
        """
        self.log(f"PEŁNA AKTUALIZACJA: START")
        
        firmware_file = self.firmware_path.get()
        
        # Flagi kontrolujące potrzebę restartu/update
        ss_updated = False
        fw_needed = False
        tz_updated = False
        
        try:
            # UŻYJ CONTEXT MANAGERA dla bezpiecznego SSH/SFTP
            with self.ssh_connection(device) as (ssh, sftp):
                
                # 1. Wykryj model PLC
                device.plc_model = self.detect_plc_model(ssh)
                
                if not device.plc_model:
                    raise Exception("Nie można wykryć modelu sterownika!")
                
                # 2. Walidacja kompatybilności firmware
                is_compatible, compat_msg = self.validate_firmware_compatibility(device, firmware_file)
                self.log(f"  {compat_msg}")
                
                if not is_compatible:
                    raise FatalUpdateError(f"{compat_msg}\n\nZATRZYMANO AKTUALIZACJĘ!")
                
                # 3. Odczyt wstępny danych
                self.log("  Wstępny odczyt danych...")
                
                # Firmware version
                stdin, stdout, stderr = ssh.exec_command("grep Arpversion /etc/plcnext/arpversion")
                fw_output = stdout.read().decode().strip()
                
                self.log(f"  Surowy output wersji firmware: '{fw_output}'")
                
                version_string = "?"
                if fw_output:
                    fw_output = fw_output.replace('Arpversion', '').strip()
                    
                    if ":" in fw_output:
                        parts = fw_output.split(':', 1) 
                        version_string = parts[1].strip() if len(parts) > 1 else "?"
                    elif "=" in fw_output:
                        version_string = fw_output.split("=")[-1].strip()
                    else:
                        version_string = fw_output.strip()
                    
                    self.log(f"  Sparsowana wersja: '{version_string}'")
                
                if version_string and version_string != "?" and version_string[0].isdigit():
                    device.firmware_version = version_string
                else:
                    device.firmware_version = "?"
                    self.log(f"  UWAGA: Nie można odczytać poprawnej wersji firmware!")
                
                # Timezone
                stdin, stdout, stderr = ssh.exec_command("cat /etc/timezone")
                device.timezone = stdout.read().decode(errors="ignore").strip()
                
                # System Services
                try:
                    remote_path = "/opt/plcnext/config/System/Scm/Default.scm.config"
                    remote_stat = sftp.stat(remote_path)
                    local_file = resource_path(SYSTEM_SERVICES_FILE)
                    if os.path.exists(local_file):
                        local_size = os.path.getsize(local_file)
                        remote_size = remote_stat.st_size
                        device.system_services_ok = "OK" if local_size == remote_size else "Różnica"
                        if device.system_services_ok == "Różnica":
                            self.log(f"  UWAGA: System Services - różnica rozmiaru: lokalny={local_size}, zdalny={remote_size}")
                    else:
                        device.system_services_ok = "Istnieje"
                except FileNotFoundError:
                    device.system_services_ok = "Brak"
                except Exception as e:
                    device.system_services_ok = "Błąd"
                    self.log(f"  UWAGA: Błąd sprawdzania System Services: {str(e)}")
                
                self.log(f"  Status System Services: {device.system_services_ok}")
                self.log(f"  Aktualna wersja FW: {device.firmware_version}")
                self.log(f"  Aktualna strefa czasowa: {device.timezone}")
                
                # 4. System Services - TYLKO UPLOAD, REBOOT PÓŹNIEJ
                if device.system_services_ok != "OK":
                    self.log(f"  System Services: {device.system_services_ok}. Wymagana aktualizacja.")
                    
                    local_sys_file = resource_path(SYSTEM_SERVICES_FILE)
                    if not os.path.exists(local_sys_file):
                        raise FatalUpdateError(f"Plik {SYSTEM_SERVICES_FILE} nie istnieje lokalnie!")
                    
                    remote_sys_path = "/opt/plcnext/config/System/Scm/Default.scm.config"
                    filename = os.path.basename(local_sys_file)
                    
                    self.log(f"  Wysyłanie {filename}...")
                    
                    self.upload_file_with_resume(
                        sftp,
                        local_sys_file,
                        remote_sys_path,
                        device=device
                    )
                    
                    self.reset_upload_progress()
                    device.system_services_ok = "OK"
                    ss_updated = True
                    self.log(f"  System Services wysłane i zweryfikowane")
                else:
                    self.log("  System Services OK - pomijam wysyłkę")
                
                # 5. Firmware - TYLKO UPLOAD, UPDATE PÓŹNIEJ
                if not self.compare_firmware_versions(device.firmware_version, firmware_file):
                    fw_needed = True
                    target_fw_version = self.get_target_fw_version(firmware_file)
                    self.log(f"  Firmware nieaktualne. Aktualna: {device.firmware_version}, Docelowa: {target_fw_version}")
                    
                    self.log("  Wysyłanie Firmware...")
                    filename = os.path.basename(firmware_file)
                    remote_fw_path = f"/opt/plcnext/{filename}"
                    
                    file_size = os.path.getsize(firmware_file)
                    
                    self.upload_file_with_resume(
                        sftp,
                        firmware_file,
                        remote_fw_path,
                        device=device
                    )
                    
                    self.reset_upload_progress()
                    self.log(f"  Plik firmware wysłany i zweryfikowany ({file_size/1024/1024:.1f} MB)")
                else:
                    self.log(f"  Firmware (v.{device.firmware_version}) jest aktualne - pomijam wysyłkę")

                # 6. Timezone - TYLKO USTAWIENIE, REBOOT PÓŹNIEJ
                if device.timezone.strip() != TIMEZONE.strip():
                    self.log(f"  Strefa czasowa niepoprawna. Ustawianie na {TIMEZONE}...")
                    
                    stdin, stdout, stderr = ssh.exec_command(
                        f"sudo sh -c 'echo {TIMEZONE} > /etc/timezone'", 
                        get_pty=True
                    )
                    stdin.write(device.password + "\n")
                    stdin.flush()
                    time.sleep(1)
                    
                    stdin, stdout, stderr = ssh.exec_command(
                        f"sudo timedatectl set-timezone {TIMEZONE}", 
                        get_pty=True
                    )
                    stdin.write(device.password + "\n")
                    stdin.flush()
                    time.sleep(1)
                    
                    device.timezone = TIMEZONE
                    tz_updated = True
                    self.log("  Strefa czasowa ustawiona")
                else:
                    self.log("  Strefa czasowa OK - pomijam zmianę")
                
                self.log("  Wszystkie transfery zakończone")
            
            # Context manager zamknął SSH tutaj - wszystkie transfery zakończone!
            
            # 7. TERAZ WYKONAJ UPDATE/REBOOT (nowe połączenie SSH)
            needs_reboot = ss_updated or tz_updated
            
            if fw_needed or needs_reboot:
                self.log("  WYKONYWANIE AKTUALIZACJI / RESTART...")
                
                if fw_needed:
                    # Firmware update - to robi automatyczny reboot
                    self.execute_firmware_update(device)
                    
                elif needs_reboot:
                    # Tylko reboot (SS lub TZ się zmieniły, ale nie FW)
                    self.execute_reboot(device)
            else:
                self.log("  INFO: Wszystkie komponenty aktualne. Pomijam restart")

            device.last_update = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            return True
            
        except Exception as e:
            self.reset_upload_progress()
            raise e


    def set_timezone_ssh(self, ssh, password):
        """
        Ustawia strefę czasową przez SSH (bez restartu).
        Używa shell interaktywnego z sudo i su.
        """
        shell = ssh.invoke_shell()
        
        def send_cmd(cmd, wait=1):
            shell.send(cmd + "\n")
            time.sleep(wait)
        
        # Ustaw hasło root
        send_cmd("sudo passwd root")
        send_cmd(password)  # sudo password
        send_cmd(ROOT_PASS)  # nowe hasło root
        send_cmd(ROOT_PASS)  # potwierdzenie
        
        # Przełącz na root
        send_cmd("su")
        send_cmd(ROOT_PASS)
        
        # Ustaw strefę czasową
        send_cmd(f"ln -sf /usr/share/zoneinfo/{TIMEZONE} /etc/localtime")
        send_cmd(f"echo '{TIMEZONE}' > /etc/timezone")
        
        # Wyłącz hasło root
        send_cmd("passwd -dl root")
        send_cmd("exit")
        
        time.sleep(2)

    def update_device_row(self, device):
        """Aktualizuje widok tabeli po zmianie statusu urządzenia."""
        self.refresh_device_tree()
        self.device_tree.update_idletasks()

    def stop_processing(self):
        """Zatrzymuje przetwarzanie."""
        if messagebox.askyesno("Potwierdzenie", "Czy na pewno chcesz zatrzymać operację?"):
            self.processing = False
            self.log("Żądanie zatrzymania operacji...")

    def log(self, message):
        """Dodaje wiadomość do kolejki logów."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_queue.put(f"[{timestamp}] {message}")

    def update_logs(self):
        """Aktualizuje okno logów z kolejki."""
        try:
            while True:
                message = self.log_queue.get_nowait()
                self.log_text.insert(tk.END, message + "\n")
                self.log_text.see(tk.END)
        except queue.Empty:
            pass
        finally:
            self.after(100, self.update_logs)

    def clear_logs(self):
        """Czyści okno logów."""
        self.log_text.delete(1.0, tk.END)

    # ============================================================================
    # RĘCZNA OBSŁUGA - pojedyncze operacje
    # ============================================================================

    def manual_read_plc(self):
        """Ręczny odczyt pojedynczego PLC."""
        ip = self.ip_entry.get()
        password = self.password_entry.get()
        if not ip or not password:
            messagebox.showerror("Błąd", "Podaj IP i hasło!")
            return
        
        device = PLCDevice("Manual", ip, password)
        threading.Thread(target=self.manual_read_worker, args=(device,), daemon=True).start()

    def manual_read_worker(self, device):
        """Worker dla ręcznego odczytu."""
        try:
            self.status_bar.config(text="Łączenie z PLC...")
            self.read_single_device(device)
            
            display_text = (
                f"Adres IP: {device.ip}\n"
                f"Aktualny czas: {device.last_check}\n"
                f"Strefa czasowa: {device.timezone}\n\n"
                f"Wersja Firmware: {device.firmware_version}\n\n"
                f"System Services: {device.system_services_ok}"
            )
            
            self.manual_data_label.config(text=display_text)
            self.status_bar.config(text="Gotowy")
            self.log(f"Odczytano dane z {device.ip}")
            
        except Exception as e:
            self.status_bar.config(text="Błąd")
            self.manual_data_label.config(text=f"Błąd odczytu:\n{str(e)}")
            self.log(f"Błąd odczytu z {device.ip}: {str(e)}")
            messagebox.showerror("Błąd", f"Błąd odczytu:\n{str(e)}")

    def select_manual_firmware(self):
        """Wybór pliku firmware dla ręcznej obsługi."""
        filepath = filedialog.askopenfilename(title="Wybierz plik firmware")
        if filepath:
            self.manual_firmware_path.set(filepath)

    def manual_set_timezone(self):
        """Ręczne ustawienie strefy czasowej."""
        ip = self.ip_entry.get()
        password = self.password_entry.get()
        if not ip or not password:
            messagebox.showerror("Błąd", "Podaj IP i hasło!")
            return
        
        response = messagebox.askyesno(
            "Potwierdzenie",
            f"Czy ustawić strefę czasową na {TIMEZONE}?\n"
            "Sterownik zostanie zrestartowany!"
        )
        if not response:
            return
        
        device = PLCDevice("Manual", ip, password)
        threading.Thread(target=self.manual_timezone_worker, args=(device,), daemon=True).start()

    def manual_timezone_worker(self, device):
        """Worker dla ustawiania strefy czasowej."""
        try:
            self.status_bar.config(text="Ustawianie strefy czasowej...")
            self.update_timezone_only(device)
            
            self.status_bar.config(text="Gotowy")
            self.after(0, lambda: messagebox.showinfo(
                "Sukces",
                f"Strefa czasowa została zmieniona na {TIMEZONE}\n"
                "Sterownik został zrestartowany."
            ))
            
        except Exception as e:
            self.status_bar.config(text="Błąd")
            self.log(f"Błąd ustawiania strefy czasowej: {str(e)}")
            self.after(0, lambda: messagebox.showerror("Błąd", f"Błąd:\n{str(e)}"))

    def manual_upload_system_services(self):
        """Ręczne wysłanie System Services."""
        ip = self.ip_entry.get()
        password = self.password_entry.get()
        if not ip or not password:
            messagebox.showerror("Błąd", "Podaj IP i hasło!")
            return
        
        local_file = resource_path(SYSTEM_SERVICES_FILE)
        if not os.path.exists(local_file):
            messagebox.showerror("Błąd", f"Plik {SYSTEM_SERVICES_FILE} nie istnieje!")
            return
        
        response = messagebox.askyesno(
            "Potwierdzenie",
            "Czy wysłać plik System Services?\n"
            "Sterownik zostanie zrestartowany!"
        )
        if not response:
            return
        
        device = PLCDevice("Manual", ip, password)
        threading.Thread(target=self.manual_sys_services_worker, args=(device,), daemon=True).start()

    def manual_sys_services_worker(self, device):
        """Worker dla wysyłania System Services."""
        try:
            self.status_bar.config(text="Wysyłanie System Services...")
            self.update_system_services_only(device)
            
            self.status_bar.config(text="Gotowy")
            self.after(0, lambda: messagebox.showinfo(
                "Sukces",
                "Plik System Services został przesłany!\n"
                "Sterownik został zrestartowany."
            ))
            
        except Exception as e:
            self.status_bar.config(text="Błąd")
            self.log(f"Błąd wysyłania System Services: {str(e)}")
            self.after(0, lambda: messagebox.showerror("Błąd", f"Błąd:\n{str(e)}"))

    def manual_upload_firmware(self):
        """Ręczne wysłanie firmware (bez wykonania update)."""
        ip = self.ip_entry.get()
        password = self.password_entry.get()
        firmware_file = self.manual_firmware_path.get()
        
        if not ip or not password:
            messagebox.showerror("Błąd", "Podaj IP i hasło!")
            return
        
        if not firmware_file or not os.path.exists(firmware_file):
            messagebox.showerror("Błąd", "Wybierz prawidłowy plik firmware!")
            return
        
        threading.Thread(target=self.manual_upload_fw_worker, 
                        args=(ip, password, firmware_file), daemon=True).start()

    def manual_upload_fw_worker(self, ip, password, firmware_file):
        """Worker dla wysyłania firmware."""
        ssh = None
        sftp = None
        try:
            self.status_bar.config(text="Wysyłanie firmware...")
            self.log(f"Łączenie z {ip} - wysyłanie firmware...")
            
            ssh = self.create_ssh_client(ip, password)

            transport = ssh.get_transport()
            if transport:
                transport.set_keepalive(self.ssh_keepalive)
            
            sftp = ssh.open_sftp()
            filename = os.path.basename(firmware_file)
            remote_path = f"/opt/plcnext/{filename}"
            
            file_size = os.path.getsize(firmware_file)
            self.log(f"Wysyłanie {filename} ({file_size/1024/1024:.1f} MB)...")
            
            self.upload_file_with_resume(sftp, firmware_file, remote_path)
            
            # Weryfikacja
            remote_size = sftp.stat(remote_path).st_size
            sftp.close()
            time.sleep(1)
            ssh.close()
            time.sleep(1)
            
            if remote_size == file_size:
                self.status_bar.config(text="Gotowy")
                self.log(f"Firmware przesłane pomyślnie")
                self.after(0, lambda: messagebox.showinfo(
                    "Sukces",
                    f"Firmware zostało przesłane!\n"
                    f"Ścieżka: {remote_path}\n"
                    f"Rozmiar: {remote_size/1024/1024:.1f} MB\n\n"
                    f"Użyj przycisku 'Wykonaj aktualizację' aby zainstalować."
                ))
            else:
                raise Exception(f"Transfer niepełny! Oczekiwano {file_size}, otrzymano {remote_size}")
            
        except Exception as e:
            if sftp:
                sftp.close()
                time.sleep(1)
            if ssh:
                ssh.close()
                time.sleep(1)
            self.status_bar.config(text="Błąd")
            self.log(f"Błąd wysyłania firmware: {str(e)}")
            self.after(0, lambda: messagebox.showerror("Błąd", f"Błąd:\n{str(e)}"))

    def manual_execute_update(self):
        """Ręczne wykonanie aktualizacji firmware."""
        ip = self.ip_entry.get()
        password = self.password_entry.get()
        
        if not ip or not password:
            messagebox.showerror("Błąd", "Podaj IP i hasło!")
            return
        
        plc_type = self.manual_plc_type_var.get()
        response = messagebox.askyesno(
            "Potwierdzenie",
            f"Czy wykonać aktualizację firmware?\n"
            f"Komenda: sudo update-axcf{plc_type}\n\n"
            "Sterownik zostanie zrestartowany!"
        )
        if not response:
            return
        
        threading.Thread(target=self.manual_execute_update_worker, 
                        args=(ip, password, plc_type), daemon=True).start()

    def manual_execute_update_worker(self, ip, password, plc_type):
        """Worker dla wykonania aktualizacji."""
        try:
            self.status_bar.config(text="Wykonywanie aktualizacji...")
            self.log(f"Łączenie z {ip} - wykonywanie aktualizacji firmware...")
            
            ssh = self.create_ssh_client(ip, password)

            transport = ssh.get_transport()
            if transport:
                transport.set_keepalive(self.ssh_keepalive)
            
            self.log(f"Wykonywanie: sudo update-axcf{plc_type}")
            stdin, stdout, stderr = ssh.exec_command(f"sudo update-axcf{plc_type}", get_pty=True)
            stdin.write(password + "\n")
            stdin.flush()
            
            output = ""
            while True:
                if stdout.channel.recv_ready():
                    chunk = stdout.read(1024).decode(errors="ignore")
                    output += chunk
                if stdout.channel.exit_status_ready():
                    break
                time.sleep(0.5)
            
            errors = stderr.read().decode(errors="ignore")
            
            ssh.close()
            time.sleep(1)
            
            if "error" in output.lower() or "failed" in output.lower() or errors.strip():
                raise Exception(f"Update zwrócił błąd:\n{output}\n{errors}")
            
            self.status_bar.config(text="Gotowy")
            self.log(f"Aktualizacja zakończona - sterownik restartuje się")
            self.after(0, lambda: messagebox.showinfo(
                "Sukces",
                "Aktualizacja firmware zakończona!\n"
                "Sterownik został zrestartowany.\n\n"
                f"Output:\n{output[:300]}..."
            ))
            
        except Exception as e:
            self.status_bar.config(text="Błąd")
            self.log(f"Błąd aktualizacji: {str(e)}")
            self.after(0, lambda: messagebox.showerror("Błąd", f"Błąd:\n{str(e)}"))


if __name__ == "__main__":
    app = BatchProcessorApp()
    app.mainloop()